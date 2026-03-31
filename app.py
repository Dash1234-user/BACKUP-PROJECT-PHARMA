"""
PharmaCare Pro — Flask + SQLite Backend  (Partition Edition)
============================================================
Drop this file in the SAME folder as index.html, app.js, styles.css.

Install & Run:
    pip install flask flask-cors
    python app.py

Then open:  http://localhost:5000

═══════════════════════════════════════════════════════════════
  DATABASE PARTITION ARCHITECTURE
═══════════════════════════════════════════════════════════════

Single pharmacare.db file with internal partitioning:

  ┌─ pharmacare.db ─────────────────────────────────────────┐
  │                                                          │
  │  settings         ← global (shared)                     │
  │  categories       ← global (shared across all modes)    │
  │  dashboard_resets ← per partition                       │
  │                                                          │
  │  ── MEDICINE DB ─────────────────────────────────────── │
  │  products    partition: 'wholesale' | 'retail' | 'both' │
  │  stock_ins   partition: 'wholesale' | 'retail' | 'both' │
  │                                                          │
  │  ── SALES HISTORY DB ──────────────────────────────── │
  │  bills       bill_store_type: 'wholesale' | 'retail'    │
  │  bill_items  (child of bills)                           │
  │                                                          │
  │  ── CREDIT DBs ─────────────────────────────────────── │
  │  credits       partition: 'wholesale' | 'both'          │
  │                (Wholesale mode: retailers who owe WS)   │
  │  shop_credits  partition: 'retail'    | 'both'          │
  │                (Retail mode: what shop owes suppliers)  │
  │                                                          │
  └──────────────────────────────────────────────────────────┘

Partition key mapping
─────────────────────────────────────────────────────────────
  Store Type              │  Partition Key
  ────────────────────────┼──────────────
  Wholesale Pharma        │  'wholesale'
  Retail Pharmacy         │  'retail'
  Hospital Pharmacy       │  'retail'
  Medical Store           │  'retail'
  Ayurvedic Store         │  'retail'

  partition = 'both'  → seed / migrated data visible to ALL modes
  partition = 'wholesale' → only visible in Wholesale Pharma mode
  partition = 'retail'    → only visible in all retail modes
"""

from flask import Flask, request, jsonify, abort, send_from_directory
from flask_cors import CORS
import sqlite3, uuid, os
from datetime import date, timedelta
from functools import wraps

# ─────────────────────────────────────────────────────────────
# APP SETUP
# ─────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.join(BASE_DIR, 'pharmacare.db')

app = Flask(__name__)
CORS(app)

# ── Serve frontend static files ───────────────────────────────
@app.route('/')
def serve_index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/<path:filename>')
def serve_static(filename):
    if filename.startswith('api/'):
        abort(404)
    return send_from_directory(BASE_DIR, filename)


# ─────────────────────────────────────────────────────────────
# PARTITION CONSTANTS & HELPERS
# ─────────────────────────────────────────────────────────────
WHOLESALE_TYPE  = 'Wholesale Pharma'
RETAIL_TYPES    = {'Retail Pharmacy', 'Hospital Pharmacy', 'Medical Store', 'Ayurvedic Store'}
PARTITION_BOTH  = 'both'
PARTITION_WS    = 'wholesale'
PARTITION_RT    = 'retail'

def _store_partition(store_type: str) -> str:
    """Map settings.store_type → partition key ('wholesale' | 'retail')."""
    return PARTITION_WS if (store_type or '').strip() == WHOLESALE_TYPE else PARTITION_RT

def _current_partition(conn) -> str:
    """Read store_type from settings and return the active partition key."""
    s = _get_settings(conn)
    return _store_partition(s.get('store_type', 'Retail Pharmacy'))

def _partition_where(partition: str) -> tuple:
    """Return (WHERE clause snippet, params) that selects partition='both' OR partition=current."""
    return "partition IN (?, 'both')", [partition]


# ─────────────────────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    return conn


def init_db():
    """
    Create all tables.  Also runs lightweight migrations (ADD COLUMN)
    so existing databases are upgraded automatically on first run.
    """
    conn = get_db()
    conn.executescript("""
    -- ── Global / Shared tables ─────────────────────────────────
    CREATE TABLE IF NOT EXISTS settings (
        id                  INTEGER PRIMARY KEY DEFAULT 1,
        store_name          TEXT DEFAULT 'My Pharmacy',
        store_type          TEXT DEFAULT 'Retail Pharmacy',
        address             TEXT DEFAULT '',
        phone               TEXT DEFAULT '',
        email               TEXT DEFAULT '',
        license_no          TEXT DEFAULT '',
        gstin               TEXT DEFAULT '',
        default_gst         REAL DEFAULT 12,
        currency            TEXT DEFAULT '₹',
        low_stock_threshold INTEGER DEFAULT 10,
        expiry_alert_days   INTEGER DEFAULT 90,
        supplier_name       TEXT DEFAULT '',
        wholesaler          TEXT DEFAULT '',
        owner_name          TEXT DEFAULT '',
        wholesaler_id       TEXT DEFAULT '',
        shop_name           TEXT DEFAULT '',
        retailer_owner      TEXT DEFAULT '',
        wholesale_upi_qr    TEXT DEFAULT '',
        retail_upi_qr       TEXT DEFAULT '',
        next_bill_no        INTEGER DEFAULT 1
    );

    CREATE TABLE IF NOT EXISTS categories (
        id   TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        desc TEXT DEFAULT ''
    );

    CREATE TABLE IF NOT EXISTS dashboard_resets (
        store_type_key  TEXT PRIMARY KEY,
        reset_date      TEXT NOT NULL
    );

    -- ── MEDICINE DB ─────────────────────────────────────────────
    -- partition: 'both' = visible to all modes (seed/migrated data)
    --            'wholesale' = only visible in Wholesale Pharma mode
    --            'retail'    = only visible in retail modes
    CREATE TABLE IF NOT EXISTS products (
        id         TEXT PRIMARY KEY,
        name       TEXT NOT NULL,
        category   TEXT REFERENCES categories(id),
        unit       TEXT DEFAULT 'Tablet',
        purchase   REAL DEFAULT 0,
        sale       REAL DEFAULT 0,
        gst        REAL DEFAULT 12,
        stock      INTEGER DEFAULT 0,
        min_stock  INTEGER DEFAULT 10,
        sku        TEXT DEFAULT '',
        expiry     TEXT DEFAULT '',
        brand      TEXT DEFAULT '',
        hsn        TEXT DEFAULT '',
        desc       TEXT DEFAULT '',
        partition  TEXT DEFAULT 'both'
    );

    -- ── SALES HISTORY DB (stock side) ──────────────────────────
    CREATE TABLE IF NOT EXISTS stock_ins (
        id           TEXT PRIMARY KEY,
        date         TEXT NOT NULL,
        product_id   TEXT REFERENCES products(id),
        product_name TEXT DEFAULT '',
        qty          INTEGER DEFAULT 0,
        price        REAL DEFAULT 0,
        batch        TEXT DEFAULT '',
        expiry       TEXT DEFAULT '',
        supplier     TEXT DEFAULT '',
        invoice_no   TEXT DEFAULT '',
        notes        TEXT DEFAULT '',
        partition    TEXT DEFAULT 'both'
    );

    -- ── SALES HISTORY DB (billing side) ─────────────────────────
    -- bill_store_type acts as the partition: 'wholesale' | 'retail'
    CREATE TABLE IF NOT EXISTS bills (
        id               TEXT PRIMARY KEY,
        bill_no          TEXT DEFAULT '',
        date             TEXT DEFAULT '',
        customer         TEXT DEFAULT '',
        phone            TEXT DEFAULT '',
        doctor           TEXT DEFAULT '',
        rx               TEXT DEFAULT '',
        payment_mode     TEXT DEFAULT 'Cash',
        notes            TEXT DEFAULT '',
        subtotal         REAL DEFAULT 0,
        total_discount   REAL DEFAULT 0,
        total_gst        REAL DEFAULT 0,
        round_off        REAL DEFAULT 0,
        grand_total      REAL DEFAULT 0,
        bill_store_type  TEXT DEFAULT 'retail',
        ws_supplier      TEXT DEFAULT '',
        ws_owner         TEXT DEFAULT '',
        ws_gstin         TEXT DEFAULT '',
        shop_name        TEXT DEFAULT '',
        shopkeeper_gstin TEXT DEFAULT '',
        rt_shop          TEXT DEFAULT '',
        rt_owner         TEXT DEFAULT '',
        rt_gstin         TEXT DEFAULT '',
        rt_license       TEXT DEFAULT '',
        rt_email         TEXT DEFAULT '',
        rt_phone         TEXT DEFAULT '',
        created_at       TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS bill_items (
        id         TEXT PRIMARY KEY,
        bill_id    TEXT REFERENCES bills(id) ON DELETE CASCADE,
        product_id TEXT DEFAULT '',
        name       TEXT DEFAULT '',
        category   TEXT DEFAULT '',
        unit       TEXT DEFAULT '',
        qty        REAL DEFAULT 0,
        unit_price REAL DEFAULT 0,
        discount   REAL DEFAULT 0,
        gst_rate   REAL DEFAULT 0,
        gst_amt    REAL DEFAULT 0,
        line_total REAL DEFAULT 0
    );

    -- ── CREDIT DB — Wholesale ────────────────────────────────────
    -- Tracks retailers/shops that owe money to THIS wholesaler.
    -- partition: 'wholesale' | 'both'
    CREATE TABLE IF NOT EXISTS credits (
        id               TEXT PRIMARY KEY,
        date             TEXT DEFAULT '',
        shop_name        TEXT DEFAULT '',
        shopkeeper_name  TEXT DEFAULT '',
        phone            TEXT DEFAULT '',
        for_item         TEXT DEFAULT '',
        amount           REAL DEFAULT 0,
        method           TEXT DEFAULT 'Cash',
        status           TEXT DEFAULT 'Pending',
        partition        TEXT DEFAULT 'wholesale'
    );

    -- ── CREDIT DB — Retail ───────────────────────────────────────
    -- Tracks what THIS retail shop owes to its suppliers/wholesalers.
    -- partition: 'retail' | 'both'
    CREATE TABLE IF NOT EXISTS shop_credits (
        id                  TEXT PRIMARY KEY,
        supplier_id         TEXT DEFAULT '',
        supplier_name       TEXT DEFAULT '',
        owner_name          TEXT DEFAULT '',
        total_purchase      REAL DEFAULT 0,
        paid                REAL DEFAULT 0,
        payment_mode        TEXT DEFAULT 'Cash',
        pending             REAL DEFAULT 0,
        last_purchase_date  TEXT DEFAULT '',
        bill_date           TEXT DEFAULT '',
        status              TEXT DEFAULT 'Pending',
        partition           TEXT DEFAULT 'retail'
    );
    """)

    # ── Migrations: add partition columns to existing databases ──
    # SQLite DEFAULT 'both' means ALL pre-existing rows read as 'both',
    # which is exactly what we want (existing data is visible to all modes).
    migrations = [
        ("products",     "partition TEXT DEFAULT 'both'"),
        ("stock_ins",    "partition TEXT DEFAULT 'both'"),
        ("credits",      "partition TEXT DEFAULT 'both'"),
        ("shop_credits", "partition TEXT DEFAULT 'both'"),
    ]
    for table, col_def in migrations:
        try:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {col_def}")
        except Exception:
            pass  # Column already exists — that's fine

    conn.execute("INSERT OR IGNORE INTO settings (id) VALUES (1)")
    conn.commit()
    conn.close()


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def uid():
    return uuid.uuid4().hex[:12]

def today_str():
    return date.today().isoformat()

def row(r):
    return dict(r) if r else None

def rows(rs):
    return [dict(r) for r in rs]

def require_json(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if request.method in ('POST', 'PUT', 'PATCH') and not request.is_json:
            return jsonify({"error": "Content-Type must be application/json"}), 400
        return f(*a, **kw)
    return wrapper

def expiry_days_left(exp_month):
    if not exp_month:
        return 9999
    try:
        y, m = int(exp_month[:4]), int(exp_month[5:7])
        last = date(y, m, 1) + timedelta(days=32)
        last = last.replace(day=1) - timedelta(days=1)
        return (last - date.today()).days
    except Exception:
        return 9999

def _settings_out(s):
    return {
        "storeName":         s.get("store_name",          "My Pharmacy"),
        "storeType":         s.get("store_type",          "Retail Pharmacy"),
        "address":           s.get("address",             ""),
        "phone":             s.get("phone",               ""),
        "email":             s.get("email",               ""),
        "license":           s.get("license_no",          ""),
        "gstin":             s.get("gstin",               ""),
        "defaultGst":        s.get("default_gst",         12),
        "currency":          s.get("currency",            "₹"),
        "lowStockThreshold": s.get("low_stock_threshold", 10),
        "expiryAlertDays":   s.get("expiry_alert_days",   90),
        "supplierName":      s.get("supplier_name",       ""),
        "wholesaler":        s.get("wholesaler",          ""),
        "ownerName":         s.get("owner_name",          ""),
        "wholesalerId":      s.get("wholesaler_id",       ""),
        "shopName":          s.get("shop_name",           ""),
        "retailerOwner":     s.get("retailer_owner",      ""),
        "wholesaleUpiQr":    s.get("wholesale_upi_qr",    ""),
        "retailUpiQr":       s.get("retail_upi_qr",       ""),
        "nextBillNo":        s.get("next_bill_no",        1),
    }

def _product_out(p):
    return {
        "id":        p["id"],
        "name":      p["name"],
        "category":  p["category"] or "",
        "unit":      p["unit"]     or "Tablet",
        "purchase":  p["purchase"] or 0,
        "sale":      p["sale"]     or 0,
        "gst":       p["gst"]      or 0,
        "stock":     p["stock"]    or 0,
        "minStock":  p["min_stock"] or 10,
        "sku":       p["sku"]      or "",
        "expiry":    p["expiry"]   or "",
        "brand":     p["brand"]    or "",
        "hsn":       p["hsn"]      or "",
        "desc":      p["desc"]     or "",
        "partition": p["partition"] if "partition" in p.keys() else PARTITION_BOTH,
    }

def _bill_item_out(i):
    return {
        "id":        i["id"],
        "productId": i["product_id"] or "",
        "name":      i["name"]       or "",
        "category":  i["category"]   or "",
        "unit":      i["unit"]       or "",
        "qty":       i["qty"]        or 0,
        "unitPrice": i["unit_price"] or 0,
        "discount":  i["discount"]   or 0,
        "gstRate":   i["gst_rate"]   or 0,
        "gstAmt":    i["gst_amt"]    or 0,
        "lineTotal": i["line_total"] or 0,
    }

def _bill_out(b, conn):
    items = conn.execute(
        "SELECT * FROM bill_items WHERE bill_id=?", (b["id"],)
    ).fetchall()
    return {
        "id":              b["id"],
        "billNo":          b["bill_no"]          or "",
        "date":            b["date"]             or "",
        "customer":        b["customer"]         or "",
        "phone":           b["phone"]            or "",
        "doctor":          b["doctor"]           or "",
        "rx":              b["rx"]               or "",
        "paymentMode":     b["payment_mode"]     or "Cash",
        "notes":           b["notes"]            or "",
        "subtotal":        b["subtotal"]         or 0,
        "totalDiscount":   b["total_discount"]   or 0,
        "totalGst":        b["total_gst"]        or 0,
        "roundOff":        b["round_off"]        or 0,
        "grandTotal":      b["grand_total"]      or 0,
        "billStoreType":   b["bill_store_type"]  or "retail",
        "wsSupplier":      b["ws_supplier"]      or "",
        "wsOwner":         b["ws_owner"]         or "",
        "wsGstin":         b["ws_gstin"]         or "",
        "shopName":        b["shop_name"]        or "",
        "shopkeeperGstin": b["shopkeeper_gstin"] or "",
        "rtShop":          b["rt_shop"]          or "",
        "rtOwner":         b["rt_owner"]         or "",
        "rtGstin":         b["rt_gstin"]         or "",
        "rtLicense":       b["rt_license"]       or "",
        "rtEmail":         b["rt_email"]         or "",
        "rtPhone":         b["rt_phone"]         or "",
        "items": [_bill_item_out(i) for i in items],
    }

def _credit_out(c):
    return {
        "id":             c["id"],
        "date":           c["date"]            or "",
        "shopName":       c["shop_name"]       or "",
        "shopkeeperName": c["shopkeeper_name"] or "",
        "phone":          c["phone"]           or "",
        "forItem":        c["for_item"]        or "",
        "amount":         c["amount"]          or 0,
        "method":         c["method"]          or "Cash",
        "status":         c["status"]          or "Pending",
        "partition":      c["partition"] if "partition" in c.keys() else PARTITION_WS,
    }

def _shop_credit_out(s):
    return {
        "id":               s["id"],
        "supplierId":       s["supplier_id"]        or "",
        "supplierName":     s["supplier_name"]      or "",
        "ownerName":        s["owner_name"]         or "",
        "totalPurchase":    s["total_purchase"]     or 0,
        "paid":             s["paid"]               or 0,
        "paymentMode":      s["payment_mode"]       or "Cash",
        "pending":          s["pending"]            or 0,
        "lastPurchaseDate": s["last_purchase_date"] or "",
        "billDate":         s["bill_date"]          or "",
        "status":           s["status"]             or "Pending",
        "partition":        s["partition"] if "partition" in s.keys() else PARTITION_RT,
    }

def _calc_item(unit_price, qty, discount, gst_rate):
    """Returns (gst_amt, line_total) for one bill line."""
    line_gross = qty * unit_price
    disc_amt   = line_gross * discount / 100
    taxable    = line_gross - disc_amt
    gst_amt    = taxable * (gst_rate / 100)
    line_total = taxable + gst_amt
    return round(gst_amt, 2), round(line_total, 2)

def _calc_totals(items):
    subtotal   = sum(it["qty"] * it["unit_price"] for it in items)
    total_disc = sum(it["qty"] * it["unit_price"] * it["discount"] / 100 for it in items)
    total_gst  = sum(it["gst_amt"] for it in items)
    raw        = subtotal - total_disc + total_gst
    grand      = round(raw)
    return {
        "subtotal":       round(subtotal,   2),
        "total_discount": round(total_disc, 2),
        "total_gst":      round(total_gst,  2),
        "round_off":      round(grand - raw, 2),
        "grand_total":    grand,
    }

def _get_settings(conn):
    r = conn.execute("SELECT * FROM settings WHERE id=1").fetchone()
    return dict(r) if r else {}

def _next_bill_no(conn):
    s = _get_settings(conn)
    n = s.get("next_bill_no", 1)
    return n, str(n).zfill(4)

def _bill_type_filter(partition: str) -> str:
    """Map partition key → bill_store_type filter expression."""
    # For bills we already have bill_store_type.
    # Wholesale sees only 'wholesale', retail sees everything that is NOT 'wholesale'.
    if partition == PARTITION_WS:
        return "bill_store_type = 'wholesale'"
    return "bill_store_type != 'wholesale'"


# ═════════════════════════════════════════════════════════════
# API ROUTES
# ═════════════════════════════════════════════════════════════

# ─────────────────────────────────────────────────────────────
# SETTINGS
# ─────────────────────────────────────────────────────────────
@app.route('/api/settings', methods=['GET'])
def get_settings():
    conn = get_db()
    s    = _get_settings(conn)
    conn.close()
    return jsonify(_settings_out(s))


@app.route('/api/settings', methods=['PUT'])
@require_json
def save_settings():
    d    = request.get_json()
    conn = get_db()
    conn.execute("""
        UPDATE settings SET
          store_name=?, store_type=?, address=?, phone=?, email=?,
          license_no=?, gstin=?, default_gst=?, currency=?,
          low_stock_threshold=?, expiry_alert_days=?,
          supplier_name=?, wholesaler=?, owner_name=?, wholesaler_id=?,
          shop_name=?, retailer_owner=?,
          wholesale_upi_qr=?, retail_upi_qr=?
        WHERE id=1
    """, (
        d.get("storeName",         "My Pharmacy"),
        d.get("storeType",         "Retail Pharmacy"),
        d.get("address",           ""),
        d.get("phone",             ""),
        d.get("email",             ""),
        d.get("license",           ""),
        d.get("gstin",             ""),
        float(d.get("defaultGst",  12)),
        d.get("currency",          "₹"),
        int(d.get("lowStockThreshold", 10)),
        int(d.get("expiryAlertDays",   90)),
        d.get("supplierName",      ""),
        d.get("wholesaler",        ""),
        d.get("ownerName",         ""),
        d.get("wholesalerId",      ""),
        d.get("shopName",          ""),
        d.get("retailerOwner",     ""),
        d.get("wholesaleUpiQr",    ""),
        d.get("retailUpiQr",       ""),
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────
# CATEGORIES  (global — no partition filtering)
# ─────────────────────────────────────────────────────────────
@app.route('/api/categories', methods=['GET'])
def get_categories():
    conn = get_db()
    rs   = conn.execute("SELECT * FROM categories ORDER BY name").fetchall()
    conn.close()
    return jsonify(rows(rs))


@app.route('/api/categories', methods=['POST'])
@require_json
def add_category():
    d    = request.get_json()
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Category name is required"}), 400
    conn   = get_db()
    exists = conn.execute(
        "SELECT id FROM categories WHERE LOWER(name)=?", (name.lower(),)
    ).fetchone()
    if exists:
        conn.close()
        return jsonify({"error": "Category already exists"}), 409
    new_id = d.get("id") or uid()
    conn.execute(
        "INSERT INTO categories(id,name,desc) VALUES(?,?,?)",
        (new_id, name, d.get("desc", ""))
    )
    conn.commit()
    r = conn.execute("SELECT * FROM categories WHERE id=?", (new_id,)).fetchone()
    conn.close()
    return jsonify(row(r)), 201


@app.route('/api/categories/<cat_id>', methods=['DELETE'])
def delete_category(cat_id):
    conn  = get_db()
    count = conn.execute(
        "SELECT COUNT(*) FROM products WHERE category=?", (cat_id,)
    ).fetchone()[0]
    if count > 0:
        conn.close()
        return jsonify({"error": f"Cannot delete: {count} medicine(s) use this category"}), 409
    conn.execute("DELETE FROM categories WHERE id=?", (cat_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────
# MEDICINE DB — PRODUCTS (partition-filtered)
# ─────────────────────────────────────────────────────────────
@app.route('/api/products', methods=['GET'])
def get_products():
    """
    Returns medicines for the CURRENT partition only.
    partition='both'  → visible to all modes (seed / migrated data)
    partition='wholesale' → only in Wholesale Pharma mode
    partition='retail'    → only in retail modes
    """
    conn = get_db()
    part = _current_partition(conn)
    rs   = conn.execute(
        "SELECT * FROM products WHERE partition IN (?, 'both') ORDER BY name",
        (part,)
    ).fetchall()
    conn.close()
    return jsonify([_product_out(r) for r in rs])


@app.route('/api/products', methods=['POST'])
@require_json
def add_product():
    d    = request.get_json()
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Medicine name is required"}), 400
    new_id = d.get("id") or uid()
    conn   = get_db()
    # Use provided partition or derive from current settings
    part   = d.get("partition") or _current_partition(conn)
    conn.execute("""
        INSERT INTO products
          (id,name,category,unit,purchase,sale,gst,stock,min_stock,sku,expiry,brand,hsn,desc,partition)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        new_id, name,
        d.get("category", ""),
        d.get("unit",     "Tablet"),
        float(d.get("purchase",  0)),
        float(d.get("sale",      0)),
        float(d.get("gst",      12)),
        int(d.get("stock",       0)),
        int(d.get("minStock",   10)),
        d.get("sku",    "").strip(),
        d.get("expiry", ""),
        d.get("brand",  "").strip(),
        d.get("hsn",    "").strip(),
        d.get("desc",   "").strip(),
        part,
    ))
    conn.commit()
    r = conn.execute("SELECT * FROM products WHERE id=?", (new_id,)).fetchone()
    conn.close()
    return jsonify(_product_out(r)), 201


@app.route('/api/products/<prod_id>', methods=['PUT'])
@require_json
def update_product(prod_id):
    d    = request.get_json()
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Medicine name is required"}), 400
    conn = get_db()
    conn.execute("""
        UPDATE products SET
          name=?, category=?, unit=?, purchase=?, sale=?, gst=?,
          stock=?, min_stock=?, sku=?, expiry=?, brand=?, hsn=?, desc=?
        WHERE id=?
    """, (
        name,
        d.get("category", ""),
        d.get("unit",     "Tablet"),
        float(d.get("purchase",  0)),
        float(d.get("sale",      0)),
        float(d.get("gst",      12)),
        int(d.get("stock",       0)),
        int(d.get("minStock",   10)),
        d.get("sku",    "").strip(),
        d.get("expiry", ""),
        d.get("brand",  "").strip(),
        d.get("hsn",    "").strip(),
        d.get("desc",   "").strip(),
        prod_id,
    ))
    conn.commit()
    r = conn.execute("SELECT * FROM products WHERE id=?", (prod_id,)).fetchone()
    conn.close()
    if not r:
        return jsonify({"error": "Product not found"}), 404
    return jsonify(_product_out(r))


@app.route('/api/products/<prod_id>', methods=['DELETE'])
def delete_product(prod_id):
    conn = get_db()
    conn.execute("DELETE FROM products WHERE id=?", (prod_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route('/api/products/<prod_id>/stock', methods=['PATCH'])
@require_json
def adjust_stock(prod_id):
    """mode = 'add' | 'remove' | 'set',  qty = integer."""
    d    = request.get_json()
    mode = d.get("mode", "add")
    qty  = int(d.get("qty", 0))
    conn = get_db()
    p    = conn.execute("SELECT stock FROM products WHERE id=?", (prod_id,)).fetchone()
    if not p:
        conn.close()
        return jsonify({"error": "Product not found"}), 404
    current   = p["stock"]
    new_stock = (current + qty if mode == "add"
                 else max(0, current - qty) if mode == "remove"
                 else max(0, qty))
    conn.execute("UPDATE products SET stock=? WHERE id=?", (new_stock, prod_id))
    conn.commit()
    r = conn.execute("SELECT * FROM products WHERE id=?", (prod_id,)).fetchone()
    conn.close()
    return jsonify(_product_out(r))


# ─────────────────────────────────────────────────────────────
# MEDICINE DB — STOCK-IN (partition-filtered)
# ─────────────────────────────────────────────────────────────
@app.route('/api/stock-ins', methods=['GET'])
def get_stock_ins():
    conn = get_db()
    part = _current_partition(conn)
    rs   = conn.execute(
        "SELECT * FROM stock_ins WHERE partition IN (?, 'both') ORDER BY date DESC, rowid DESC LIMIT 200",
        (part,)
    ).fetchall()
    conn.close()
    return jsonify([{
        "id":          r["id"],
        "date":        r["date"],
        "productId":   r["product_id"],
        "productName": r["product_name"],
        "qty":         r["qty"],
        "price":       r["price"],
        "batch":       r["batch"],
        "expiry":      r["expiry"],
        "supplier":    r["supplier"],
        "invoiceNo":   r["invoice_no"],
        "notes":       r["notes"],
        "partition":   r["partition"] if "partition" in r.keys() else PARTITION_BOTH,
    } for r in rs])


@app.route('/api/stock-ins', methods=['POST'])
@require_json
def add_stock_in():
    d       = request.get_json()
    prod_id = d.get("productId", "")
    qty     = int(d.get("qty", 0))
    price   = float(d.get("price", 0))
    batch   = d.get("batch",  "").strip()
    expiry  = d.get("expiry", "")

    if not prod_id or qty < 1:
        return jsonify({"error": "Product and valid quantity are required"}), 400

    conn   = get_db()
    new_id = d.get("id") or uid()
    part   = d.get("partition") or _current_partition(conn)

    conn.execute("""
        INSERT INTO stock_ins
          (id,date,product_id,product_name,qty,price,batch,expiry,supplier,invoice_no,notes,partition)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        new_id,
        d.get("date") or today_str(),
        prod_id,
        d.get("productName", ""),
        qty, price, batch, expiry,
        d.get("supplier",  "").strip(),
        d.get("invoiceNo", "").strip(),
        d.get("notes",     "").strip(),
        part,
    ))

    # Update product stock, purchase price, batch, expiry
    p = conn.execute("SELECT * FROM products WHERE id=?", (prod_id,)).fetchone()
    if p:
        new_stock = p["stock"] + qty
        new_price = price    if price  > 0 else p["purchase"]
        new_sku   = batch    if batch        else p["sku"]
        new_exp   = expiry   if expiry       else p["expiry"]
        conn.execute(
            "UPDATE products SET stock=?, purchase=?, sku=?, expiry=? WHERE id=?",
            (new_stock, new_price, new_sku, new_exp, prod_id)
        )

    conn.commit()
    updated = conn.execute("SELECT * FROM products WHERE id=?", (prod_id,)).fetchone()
    conn.close()
    return jsonify({
        "ok": True,
        "updatedProduct": _product_out(updated) if updated else None,
    }), 201


# ─────────────────────────────────────────────────────────────
# BILLS — calculate (no DB write)
# ─────────────────────────────────────────────────────────────
@app.route('/api/bills/calculate', methods=['POST'])
@require_json
def calculate_bill():
    d         = request.get_json()
    raw_items = d.get("items", [])
    out_items = []
    for it in raw_items:
        qty        = float(it.get("qty",       1))
        unit_price = float(it.get("unitPrice", 0))
        discount   = float(it.get("discount",  0))
        gst_rate   = float(it.get("gstRate",   0))
        gst_amt, line_total = _calc_item(unit_price, qty, discount, gst_rate)
        out_items.append({
            **it,
            "qty":       qty,
            "unitPrice": unit_price,
            "discount":  discount,
            "gstRate":   gst_rate,
            "gstAmt":    gst_amt,
            "lineTotal": line_total,
        })
    totals = _calc_totals([{
        "qty":        it["qty"],
        "unit_price": it["unitPrice"],
        "discount":   it["discount"],
        "gst_amt":    it["gstAmt"],
    } for it in out_items])
    return jsonify({"items": out_items, "totals": totals})


# ─────────────────────────────────────────────────────────────
# BILLS — next number
# ─────────────────────────────────────────────────────────────
@app.route('/api/bills/next-number', methods=['GET'])
def next_bill_number():
    conn   = get_db()
    n, fmt = _next_bill_no(conn)
    s      = _get_settings(conn)
    conn.close()
    return jsonify({"nextBillNo": n, "formatted": fmt, "gstin": s.get("gstin", "")})


# ─────────────────────────────────────────────────────────────
# SALES HISTORY DB — BILLS list & get (partition-filtered)
# ─────────────────────────────────────────────────────────────
@app.route('/api/bills', methods=['GET'])
def get_bills():
    conn   = get_db()
    part   = _current_partition(conn)
    bill_w = _bill_type_filter(part)

    query  = f"SELECT * FROM bills WHERE {bill_w}"
    params = []

    q     = request.args.get("q",       "")
    frm   = request.args.get("from",    "")
    to    = request.args.get("to",      "")
    pay   = request.args.get("payment", "")
    btype = request.args.get("type",    "")
    limit = int(request.args.get("limit", 500))

    if q:
        query  += " AND (bill_no LIKE ? OR customer LIKE ? OR doctor LIKE ? OR phone LIKE ?)"
        p       = f"%{q}%"
        params += [p, p, p, p]
    if frm:
        query  += " AND date >= ?"; params.append(frm)
    if to:
        query  += " AND date <= ?"; params.append(to)
    if pay:
        query  += " AND payment_mode = ?"; params.append(pay)
    if btype:
        query  += " AND bill_store_type = ?"; params.append(btype)

    query  += " ORDER BY created_at DESC LIMIT ?"
    params.append(limit)

    bill_rows = conn.execute(query, params).fetchall()
    result    = [_bill_out(b, conn) for b in bill_rows]
    conn.close()
    return jsonify(result)


@app.route('/api/bills/<bill_id>', methods=['GET'])
def get_bill(bill_id):
    conn = get_db()
    b    = conn.execute("SELECT * FROM bills WHERE id=?", (bill_id,)).fetchone()
    if not b:
        conn.close()
        return jsonify({"error": "Bill not found"}), 404
    result = _bill_out(b, conn)
    conn.close()
    return jsonify(result)


# ─────────────────────────────────────────────────────────────
# SALES HISTORY DB — BILLS create
# ─────────────────────────────────────────────────────────────
@app.route('/api/bills', methods=['POST'])
@require_json
def create_bill():
    d         = request.get_json()
    conn      = get_db()
    s         = _get_settings(conn)
    raw_items = d.get("items", [])

    if not raw_items:
        conn.close()
        return jsonify({"error": "Bill must have at least one item"}), 400

    calc_items = []
    for it in raw_items:
        qty        = float(it.get("qty",       1))
        unit_price = float(it.get("unitPrice", 0))
        discount   = float(it.get("discount",  0))
        gst_rate   = float(it.get("gstRate",   0))
        gst_amt, line_total = _calc_item(unit_price, qty, discount, gst_rate)
        calc_items.append({
            "productId": it.get("productId", ""),
            "name":      it.get("name",      ""),
            "category":  it.get("category",  ""),
            "unit":      it.get("unit",       ""),
            "qty":       qty,
            "unitPrice": unit_price,
            "discount":  discount,
            "gstRate":   gst_rate,
            "gstAmt":    gst_amt,
            "lineTotal": line_total,
        })

    totals = _calc_totals([{
        "qty":        it["qty"],
        "unit_price": it["unitPrice"],
        "discount":   it["discount"],
        "gst_amt":    it["gstAmt"],
    } for it in calc_items])

    n, fmt   = _next_bill_no(conn)
    is_ws    = d.get("billStoreType", "retail") == "wholesale"
    ws_gstin = d.get("wsGstin", "") or s.get("gstin", "")
    bill_no  = f"{ws_gstin}-{fmt}" if is_ws and ws_gstin else fmt

    bill_id = d.get("id") or uid()
    conn.execute("""
        INSERT INTO bills
          (id,bill_no,date,customer,phone,doctor,rx,payment_mode,notes,
           subtotal,total_discount,total_gst,round_off,grand_total,bill_store_type,
           ws_supplier,ws_owner,ws_gstin,shop_name,shopkeeper_gstin,
           rt_shop,rt_owner,rt_gstin,rt_license,rt_email,rt_phone)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        bill_id, bill_no,
        d.get("date") or today_str(),
        d.get("customer",        "Walk-in"),
        d.get("phone",           ""),
        d.get("doctor",          ""),
        d.get("rx",              ""),
        d.get("paymentMode",     "Cash"),
        d.get("notes",           ""),
        totals["subtotal"],
        totals["total_discount"],
        totals["total_gst"],
        totals["round_off"],
        totals["grand_total"],
        "wholesale" if is_ws else "retail",
        d.get("wsSupplier",      ""),
        d.get("wsOwner",         ""),
        ws_gstin,
        d.get("shopName",        ""),
        d.get("shopkeeperGstin", ""),
        d.get("rtShop",          ""),
        d.get("rtOwner",         ""),
        d.get("rtGstin",         ""),
        d.get("rtLicense",       ""),
        d.get("rtEmail",         ""),
        d.get("rtPhone",         ""),
    ))

    low_stock_alerts = []
    for it in calc_items:
        conn.execute("""
            INSERT INTO bill_items
              (id,bill_id,product_id,name,category,unit,qty,unit_price,
               discount,gst_rate,gst_amt,line_total)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            uid(), bill_id,
            it["productId"], it["name"], it["category"], it["unit"],
            it["qty"], it["unitPrice"], it["discount"],
            it["gstRate"], it["gstAmt"], it["lineTotal"],
        ))
        if it["productId"]:
            conn.execute(
                "UPDATE products SET stock = MAX(0, stock - ?) WHERE id=?",
                (int(it["qty"]), it["productId"])
            )
            p = conn.execute(
                "SELECT name, stock, min_stock FROM products WHERE id=?",
                (it["productId"],)
            ).fetchone()
            if p and p["stock"] <= p["min_stock"]:
                low_stock_alerts.append({"name": p["name"], "stock": p["stock"]})

    conn.execute("UPDATE settings SET next_bill_no = next_bill_no + 1 WHERE id=1")
    conn.commit()

    saved  = conn.execute("SELECT * FROM bills WHERE id=?", (bill_id,)).fetchone()
    result = _bill_out(saved, conn)
    new_n  = n + 1
    conn.close()

    return jsonify({
        "bill":           result,
        "nextBillNo":     new_n,
        "lowStockAlerts": low_stock_alerts,
    }), 201


@app.route('/api/bills/<bill_id>', methods=['DELETE'])
def delete_bill(bill_id):
    conn = get_db()
    conn.execute("DELETE FROM bill_items WHERE bill_id=?", (bill_id,))
    conn.execute("DELETE FROM bills WHERE id=?",           (bill_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────
# CREDIT DB — WHOLESALE credits (partition='wholesale'|'both')
# ─────────────────────────────────────────────────────────────
@app.route('/api/credits', methods=['GET'])
def get_credits():
    """
    Returns credits for Wholesale Pharma partition only.
    These represent retailers/shops that owe money to this wholesaler.
    """
    conn = get_db()
    rs   = conn.execute(
        "SELECT * FROM credits WHERE partition IN ('wholesale', 'both') ORDER BY date DESC"
    ).fetchall()
    conn.close()
    return jsonify([_credit_out(r) for r in rs])


@app.route('/api/credits', methods=['POST'])
@require_json
def add_credit():
    d = request.get_json()
    if not d.get("shopName") or not d.get("shopkeeperName"):
        return jsonify({"error": "Shop name and shopkeeper name are required"}), 400
    new_id = d.get("id") or uid()
    conn   = get_db()
    conn.execute("""
        INSERT INTO credits
          (id,date,shop_name,shopkeeper_name,phone,for_item,amount,method,status,partition)
        VALUES(?,?,?,?,?,?,?,?,?,?)
    """, (
        new_id,
        d.get("date") or today_str(),
        d.get("shopName",       ""),
        d.get("shopkeeperName", ""),
        d.get("phone",          ""),
        d.get("forItem",        ""),
        float(d.get("amount",   0)),
        d.get("method",         "Cash"),
        d.get("status",         "Pending"),
        PARTITION_WS,   # always wholesale
    ))
    conn.commit()
    r = conn.execute("SELECT * FROM credits WHERE id=?", (new_id,)).fetchone()
    conn.close()
    return jsonify(_credit_out(r)), 201


@app.route('/api/credits/<credit_id>', methods=['PATCH'])
@require_json
def update_credit_status(credit_id):
    d    = request.get_json()
    conn = get_db()
    conn.execute(
        "UPDATE credits SET status=? WHERE id=?",
        (d.get("status", "Pending"), credit_id)
    )
    conn.commit()
    r = conn.execute("SELECT * FROM credits WHERE id=?", (credit_id,)).fetchone()
    conn.close()
    if not r:
        return jsonify({"error": "Credit not found"}), 404
    return jsonify(_credit_out(r))


@app.route('/api/credits/<credit_id>', methods=['DELETE'])
def delete_credit(credit_id):
    conn = get_db()
    conn.execute("DELETE FROM credits WHERE id=?", (credit_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────
# CREDIT DB — RETAIL shop_credits (partition='retail'|'both')
# ─────────────────────────────────────────────────────────────
@app.route('/api/shop-credits', methods=['GET'])
def get_shop_credits():
    """
    Returns shop credits for retail partition only.
    These represent what this retail shop owes to suppliers/wholesalers.
    """
    conn = get_db()
    rs   = conn.execute(
        "SELECT * FROM shop_credits WHERE partition IN ('retail', 'both') ORDER BY bill_date DESC, rowid DESC"
    ).fetchall()
    conn.close()
    return jsonify([_shop_credit_out(r) for r in rs])


@app.route('/api/shop-credits/fetch/<supplier_id>', methods=['GET'])
def fetch_shop_credit_by_supplier(supplier_id):
    conn = get_db()
    r    = conn.execute("""
        SELECT * FROM shop_credits
        WHERE LOWER(supplier_id)=? AND partition IN ('retail', 'both')
        ORDER BY bill_date DESC, rowid DESC
        LIMIT 1
    """, (supplier_id.lower(),)).fetchone()
    conn.close()
    if not r:
        return jsonify({"error": "No record found"}), 404
    return jsonify(_shop_credit_out(r))


@app.route('/api/shop-credits', methods=['POST'])
@require_json
def add_shop_credit():
    d = request.get_json()
    if not d.get("supplierId") or not d.get("supplierName"):
        return jsonify({"error": "Supplier ID and name are required"}), 400
    new_id = d.get("id") or uid()
    conn   = get_db()
    conn.execute("""
        INSERT INTO shop_credits
          (id,supplier_id,supplier_name,owner_name,total_purchase,paid,
           payment_mode,pending,last_purchase_date,bill_date,status,partition)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        new_id,
        d.get("supplierId",       ""),
        d.get("supplierName",     ""),
        d.get("ownerName",        ""),
        float(d.get("totalPurchase", 0)),
        float(d.get("paid",          0)),
        d.get("paymentMode",      "Cash"),
        float(d.get("pending",       0)),
        d.get("lastPurchaseDate") or today_str(),
        d.get("billDate")         or today_str(),
        d.get("status",           "Pending"),
        PARTITION_RT,   # always retail
    ))
    conn.commit()
    r = conn.execute("SELECT * FROM shop_credits WHERE id=?", (new_id,)).fetchone()
    conn.close()
    return jsonify(_shop_credit_out(r)), 201


@app.route('/api/shop-credits/<sc_id>', methods=['DELETE'])
def delete_shop_credit(sc_id):
    conn = get_db()
    conn.execute("DELETE FROM shop_credits WHERE id=?", (sc_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route('/api/shop-credits/supplier/<supplier_id>/history', methods=['GET'])
def supplier_history(supplier_id):
    conn = get_db()
    rs   = conn.execute("""
        SELECT * FROM shop_credits
        WHERE LOWER(supplier_id)=? AND partition IN ('retail', 'both')
        ORDER BY bill_date DESC
    """, (supplier_id.lower(),)).fetchall()
    conn.close()
    if not rs:
        return jsonify({"error": "No records found"}), 404
    records         = [_shop_credit_out(r) for r in rs]
    total_purchased = round(sum(r["totalPurchase"] for r in records), 2)
    total_paid      = round(sum(r["paid"]          for r in records), 2)
    current_pending = records[0]["pending"] if records else 0
    return jsonify({
        "records":        records,
        "totalPurchased": total_purchased,
        "totalPaid":      total_paid,
        "currentPending": current_pending,
        "supplierName":   records[0]["supplierName"] if records else "",
        "ownerName":      records[0]["ownerName"]    if records else "",
    })


# ─────────────────────────────────────────────────────────────
# DASHBOARD (partition-aware)
# ─────────────────────────────────────────────────────────────
@app.route('/api/dashboard', methods=['GET'])
def get_dashboard():
    conn       = get_db()
    s          = _get_settings(conn)
    today      = today_str()
    alert_days = s.get("expiry_alert_days", 90)
    store_type = s.get("store_type", "Retail Pharmacy").strip()
    part       = _store_partition(store_type)
    bill_where = _bill_type_filter(part)
    bill_type_key = part  # 'wholesale' | 'retail'

    # Get reset date for this partition
    reset_row = conn.execute(
        "SELECT reset_date FROM dashboard_resets WHERE store_type_key=?",
        (bill_type_key,)
    ).fetchone()
    reset_date = reset_row["reset_date"] if reset_row else None

    date_filter = ""
    date_params_today = [today]
    if reset_date:
        date_filter = " AND date >= ?"
        date_params_today.append(reset_date)

    # Today revenue (partition-filtered)
    today_bills = conn.execute(
        f"SELECT grand_total FROM bills WHERE date=? AND {bill_where}{date_filter}",
        date_params_today
    ).fetchall()
    today_rev = sum(r["grand_total"] for r in today_bills)

    # Products — partition-filtered for stock/expiry alerts
    products      = conn.execute(
        "SELECT * FROM products WHERE partition IN (?, 'both')", (part,)
    ).fetchall()
    low_stock     = [_product_out(p) for p in products if p["stock"] <= p["min_stock"]]
    expiry_alerts, expired_count = [], 0
    for p in products:
        days = expiry_days_left(p["expiry"])
        if days < 0:
            expired_count += 1
            expiry_alerts.append({**_product_out(p), "daysLeft": days})
        elif days <= alert_days:
            expiry_alerts.append({**_product_out(p), "daysLeft": days})

    # Revenue last 7 days
    rev_7 = []
    for i in range(6, -1, -1):
        d     = (date.today() - timedelta(days=i)).isoformat()
        params = [d]
        extra  = ""
        if reset_date and d < reset_date:
            rev_7.append({"date": d, "revenue": 0.0})
            continue
        if reset_date:
            extra = " AND date >= ?"
            params.append(reset_date)
        total = conn.execute(
            f"SELECT COALESCE(SUM(grand_total),0) FROM bills WHERE date=? AND {bill_where}{extra}",
            params
        ).fetchone()[0]
        rev_7.append({"date": d, "revenue": round(float(total), 2)})

    # Top 8 products by units sold
    reset_clause = f" AND b.date >= '{reset_date}'" if reset_date else ""
    top_rows = conn.execute(f"""
        SELECT bi.name, SUM(bi.qty) as units
        FROM bill_items bi
        JOIN bills b ON b.id = bi.bill_id
        WHERE b.{bill_where}{reset_clause}
        GROUP BY bi.name ORDER BY units DESC LIMIT 8
    """).fetchall()
    top_products = [{"name": r["name"], "units": r["units"]} for r in top_rows]

    # Weekly profit (current month, partition-filtered)
    now         = date.today()
    month_start = now.replace(day=1).isoformat()
    next_mo     = (now.replace(day=28) + timedelta(days=4)).replace(day=1)
    month_end   = (next_mo - timedelta(days=1)).isoformat()
    effective_start = max(month_start, reset_date) if reset_date else month_start
    purchase_map = {p["id"]: p["purchase"] for p in products}
    bills_month  = conn.execute(
        f"SELECT id, date FROM bills WHERE date>=? AND date<=? AND {bill_where}",
        (effective_start, month_end)
    ).fetchall()
    week_profit = [0.0, 0.0, 0.0, 0.0]
    for b in bills_month:
        day_n = int(b["date"].split("-")[2])
        w_idx = 0 if day_n <= 7 else 1 if day_n <= 14 else 2 if day_n <= 21 else 3
        for it in conn.execute(
            "SELECT product_id, unit_price, qty, discount FROM bill_items WHERE bill_id=?",
            (b["id"],)
        ).fetchall():
            purchase  = purchase_map.get(it["product_id"], 0)
            sale_disc = it["unit_price"] * (1 - it["discount"] / 100)
            week_profit[w_idx] += (sale_disc - purchase) * it["qty"]
    week_profit = [round(v, 2) for v in week_profit]

    # Recent 8 bills
    reset_q = f" AND date >= '{reset_date}'" if reset_date else ""
    recent_rows  = conn.execute(
        f"SELECT * FROM bills WHERE {bill_where}{reset_q} ORDER BY created_at DESC LIMIT 8"
    ).fetchall()
    recent_bills = [_bill_out(b, conn) for b in recent_rows]

    conn.close()
    return jsonify({
        "todayRevenue":   round(today_rev, 2),
        "todayBillCount": len(today_bills),
        "totalProducts":  len(products),
        "lowStockCount":  len(low_stock),
        "lowStockItems":  low_stock[:10],
        "expiryAlerts":   expiry_alerts[:6],
        "expiredCount":   expired_count,
        "revenue7Days":   rev_7,
        "topProducts":    top_products,
        "weekProfit":     week_profit,
        "recentBills":    recent_bills,
        "products":       [_product_out(p) for p in products],
        "resetDate":      reset_date,
        "storeTypeKey":   bill_type_key,
    })


# ─────────────────────────────────────────────────────────────
# DASHBOARD RESETS
# ─────────────────────────────────────────────────────────────
@app.route('/api/dashboard/resets', methods=['GET'])
def get_dashboard_resets():
    conn = get_db()
    rs   = conn.execute("SELECT store_type_key, reset_date FROM dashboard_resets").fetchall()
    conn.close()
    return jsonify({"resets": [{"storeTypeKey": r["store_type_key"], "resetDate": r["reset_date"]} for r in rs]})


@app.route('/api/dashboard/reset', methods=['POST'])
def post_dashboard_reset():
    d     = request.get_json(force=True) or {}
    key   = d.get("storeTypeKey", "").strip()
    rdate = d.get("resetDate", "").strip()
    if key not in ("wholesale", "retail"):
        return jsonify({"error": "Invalid storeTypeKey"}), 400
    if not rdate:
        return jsonify({"error": "resetDate required"}), 400
    conn = get_db()
    conn.execute(
        "INSERT INTO dashboard_resets (store_type_key, reset_date) VALUES (?,?) "
        "ON CONFLICT(store_type_key) DO UPDATE SET reset_date=excluded.reset_date",
        (key, rdate)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "storeTypeKey": key, "resetDate": rdate})


# ─────────────────────────────────────────────────────────────
# ANALYSIS (partition-aware)
# ─────────────────────────────────────────────────────────────
@app.route('/api/analysis', methods=['GET'])
def get_analysis():
    days      = int(request.args.get("days", 7))
    from_date = (date.today() - timedelta(days=days)).isoformat()
    conn      = get_db()
    part      = _current_partition(conn)
    bill_w    = _bill_type_filter(part)

    bills     = conn.execute(
        f"SELECT * FROM bills WHERE date>=? AND {bill_w}", (from_date,)
    ).fetchall()
    bill_ids  = [b["id"] for b in bills]

    total_rev  = sum(b["grand_total"] for b in bills)
    avg_bill   = (total_rev / len(bills)) if bills else 0
    prod_sales, cat_sales, pay_totals, rev_by_day = {}, {}, {}, {}

    for b in bills:
        pm = b["payment_mode"]
        pay_totals[pm]    = round(pay_totals.get(pm, 0) + b["grand_total"], 2)
        rev_by_day[b["date"]] = round(rev_by_day.get(b["date"], 0) + b["grand_total"], 2)

    if bill_ids:
        ph    = ",".join("?" * len(bill_ids))
        items = conn.execute(
            f"SELECT * FROM bill_items WHERE bill_id IN ({ph})", bill_ids
        ).fetchall()
        cats  = {r["id"]: r["name"] for r in
                 conn.execute("SELECT id, name FROM categories").fetchall()}
        for it in items:
            n  = it["name"]
            cn = cats.get(it["category"], "Uncategorized")
            if n not in prod_sales:
                prod_sales[n] = {"units": 0, "revenue": 0, "category": it["category"]}
            prod_sales[n]["units"]   += it["qty"]
            prod_sales[n]["revenue"]  = round(prod_sales[n]["revenue"] + it["line_total"], 2)
            cat_sales[cn] = round(cat_sales.get(cn, 0) + it["line_total"], 2)

    sorted_prods = sorted(prod_sales.items(), key=lambda x: x[1]["revenue"], reverse=True)
    top_product  = sorted_prods[0][0] if sorted_prods else "—"

    rev_list = []
    for i in range(days - 1, -1, -1):
        d_str = (date.today() - timedelta(days=i)).isoformat()
        rev_list.append({"date": d_str, "revenue": rev_by_day.get(d_str, 0)})

    conn.close()
    return jsonify({
        "totalBills":       len(bills),
        "totalRevenue":     round(total_rev, 2),
        "avgBillValue":     round(avg_bill, 2),
        "topProduct":       top_product,
        "productSales":     [{"name": k, **v} for k, v in sorted_prods],
        "categorySales":    [{"name": k, "revenue": v} for k, v in
                             sorted(cat_sales.items(), key=lambda x: -x[1])],
        "paymentBreakdown": [{"mode": k, "total": v} for k, v in pay_totals.items()],
        "revenueByDay":     rev_list,
    })


# ─────────────────────────────────────────────────────────────
# EXPIRY TRACKER (partition-filtered)
# ─────────────────────────────────────────────────────────────
@app.route('/api/expiry', methods=['GET'])
def get_expiry():
    conn     = get_db()
    part     = _current_partition(conn)
    products = conn.execute(
        "SELECT * FROM products WHERE partition IN (?, 'both')", (part,)
    ).fetchall()
    conn.close()
    result   = {"expired": [], "within30": [], "within60": [], "within90": [], "safe": []}
    for p in products:
        days = expiry_days_left(p["expiry"])
        pd   = {**_product_out(p), "daysLeft": days}
        if days < 0:      result["expired"].append(pd)
        elif days <= 30:  result["within30"].append(pd)
        elif days <= 60:  result["within60"].append(pd)
        elif days <= 90:  result["within90"].append(pd)
        else:             result["safe"].append(pd)
    return jsonify({**result, "counts": {k: len(v) for k, v in result.items()}})


# ─────────────────────────────────────────────────────────────
# PARTITION INFO  — useful for debugging / frontend awareness
# ─────────────────────────────────────────────────────────────
@app.route('/api/partition-info', methods=['GET'])
def partition_info():
    """
    Returns info about which partition is active and record counts per partition.
    Useful for debugging and the settings UI.
    """
    conn = get_db()
    s    = _get_settings(conn)
    part = _store_partition(s.get('store_type', 'Retail Pharmacy'))

    def count(table, p):
        return conn.execute(
            f"SELECT COUNT(*) FROM {table} WHERE partition=?", (p,)
        ).fetchone()[0]
    def count_both(table):
        return conn.execute(
            f"SELECT COUNT(*) FROM {table} WHERE partition='both'"
        ).fetchone()[0]

    info = {
        "activePartition": part,
        "storeType":       s.get("store_type", "Retail Pharmacy"),
        "medicineCounts": {
            "shared":    count_both("products"),
            "wholesale": count("products", PARTITION_WS),
            "retail":    count("products", PARTITION_RT),
            "visible":   conn.execute(
                "SELECT COUNT(*) FROM products WHERE partition IN (?, 'both')", (part,)
            ).fetchone()[0],
        },
        "stockInCounts": {
            "shared":    count_both("stock_ins"),
            "wholesale": count("stock_ins", PARTITION_WS),
            "retail":    count("stock_ins", PARTITION_RT),
        },
        "creditCounts": {
            "wholesale_shared":   count_both("credits"),
            "wholesale_specific": count("credits", PARTITION_WS),
            "retail_shared":      count_both("shop_credits"),
            "retail_specific":    count("shop_credits", PARTITION_RT),
        },
        "billCounts": {
            "wholesale": conn.execute(
                "SELECT COUNT(*) FROM bills WHERE bill_store_type='wholesale'"
            ).fetchone()[0],
            "retail": conn.execute(
                "SELECT COUNT(*) FROM bills WHERE bill_store_type!='wholesale'"
            ).fetchone()[0],
        },
    }
    conn.close()
    return jsonify(info)


# ─────────────────────────────────────────────────────────────
# STATE — full STATE sync (partition-aware)
# ─────────────────────────────────────────────────────────────
@app.route('/api/state', methods=['GET'])
def get_state():
    """
    Return the entire app STATE filtered to the CURRENT partition.
    app.js will receive only data relevant to the active pharmacy type.
    """
    conn = get_db()
    s    = _settings_out(_get_settings(conn))
    part = _store_partition(s["storeType"])
    bill_w = _bill_type_filter(part)

    # Shared categories
    categories = [dict(r) for r in conn.execute(
        "SELECT * FROM categories ORDER BY name").fetchall()]

    # Partition-filtered products (Medicine DB)
    products = [_product_out(r) for r in conn.execute(
        "SELECT * FROM products WHERE partition IN (?, 'both') ORDER BY name",
        (part,)
    ).fetchall()]

    # Partition-filtered bills (Sales History DB)
    bills_raw = conn.execute(
        f"SELECT * FROM bills WHERE {bill_w} ORDER BY date DESC, created_at DESC"
    ).fetchall()
    bills = [_bill_out(b, conn) for b in bills_raw]

    def _si_out(r):
        return {
            "id":          r["id"],
            "date":        r["date"]         or "",
            "productId":   r["product_id"]   or "",
            "productName": r["product_name"] or "",
            "qty":         r["qty"]          or 0,
            "price":       r["price"]        or 0,
            "batch":       r["batch"]        or "",
            "expiry":      r["expiry"]       or "",
            "supplier":    r["supplier"]     or "",
            "invoiceNo":   r["invoice_no"]   or "",
            "notes":       r["notes"]        or "",
            "partition":   r["partition"] if "partition" in r.keys() else PARTITION_BOTH,
        }

    # Partition-filtered stock-ins (Medicine DB / Sales History stock side)
    stock_ins = [_si_out(r) for r in conn.execute(
        "SELECT * FROM stock_ins WHERE partition IN (?, 'both') ORDER BY date DESC",
        (part,)
    ).fetchall()]

    # Credit DB — Wholesale: visible when partition=wholesale
    credits = [_credit_out(r) for r in conn.execute(
        "SELECT * FROM credits WHERE partition IN ('wholesale', 'both') ORDER BY date DESC"
    ).fetchall()]

    # Credit DB — Retail: visible when partition=retail
    shop_credits = [_shop_credit_out(r) for r in conn.execute(
        "SELECT * FROM shop_credits WHERE partition IN ('retail', 'both') ORDER BY bill_date DESC"
    ).fetchall()]

    reset_rows = conn.execute(
        "SELECT store_type_key, reset_date FROM dashboard_resets").fetchall()
    dashboard_resets = {r["store_type_key"]: r["reset_date"] for r in reset_rows}

    conn.close()
    return jsonify({
        "settings":        s,
        "categories":      categories,
        "products":        products,
        "bills":           bills,
        "stockIns":        stock_ins,
        "credits":         credits,
        "shopCredits":     shop_credits,
        "nextBillNo":      s.get("nextBillNo", 1),
        "dashboardResets": dashboard_resets,
    })


@app.route('/api/state', methods=['POST'])
@require_json
def save_state():
    """
    Partition-aware full-replace sync — called by app.js saveState().

    Strategy per data type:
    ─────────────────────────────────────────────────────────────
    categories   : Clear all + reinsert (globally shared)
    products     : Upsert payload items, preserving existing
                   'both' partition; delete current-partition items
                   that were removed from the frontend.
    stock_ins    : Same as products
    bills        : Replace only current partition's bill_store_type bills
    credits      : Clear all + reinsert (naturally wholesale-only table)
    shop_credits : Clear all + reinsert (naturally retail-only table)
    ─────────────────────────────────────────────────────────────
    """
    data = request.get_json()
    conn = get_db()
    part = _current_partition(conn)
    bill_type = PARTITION_WS if part == PARTITION_WS else PARTITION_RT

    # ── Settings ──────────────────────────────────────────────
    s = data.get("settings", {})
    conn.execute("""
        UPDATE settings SET
          store_name=?, store_type=?, address=?, phone=?, email=?,
          license_no=?, gstin=?, default_gst=?, currency=?,
          low_stock_threshold=?, expiry_alert_days=?,
          supplier_name=?, wholesaler=?, owner_name=?, wholesaler_id=?,
          shop_name=?, retailer_owner=?,
          wholesale_upi_qr=?, retail_upi_qr=?, next_bill_no=?
        WHERE id=1
    """, (
        s.get("storeName",         "My Pharmacy"),
        s.get("storeType",         "Retail Pharmacy"),
        s.get("address",           ""),
        s.get("phone",             ""),
        s.get("email",             ""),
        s.get("license",           ""),
        s.get("gstin",             ""),
        float(s.get("defaultGst",  12)),
        s.get("currency",          "₹"),
        int(s.get("lowStockThreshold", 10)),
        int(s.get("expiryAlertDays",   90)),
        s.get("supplierName",      ""),
        s.get("wholesaler",        ""),
        s.get("ownerName",         ""),
        s.get("wholesalerId",      ""),
        s.get("shopName",          ""),
        s.get("retailerOwner",     ""),
        s.get("wholesaleUpiQr",    ""),
        s.get("retailUpiQr",       ""),
        int(data.get("nextBillNo", s.get("nextBillNo", 1))),
    ))

    # ── Categories (globally shared — full replace) ───────────
    conn.execute("DELETE FROM categories")
    for c in data.get("categories", []):
        conn.execute("INSERT INTO categories(id,name,desc) VALUES(?,?,?)",
            (c["id"], c.get("name",""), c.get("desc","")))

    # ── MEDICINE DB — Products (partition-aware upsert) ────────
    # Snapshot existing partition assignments before touching rows
    existing_prod_parts = {
        r["id"]: r["partition"]
        for r in conn.execute("SELECT id, partition FROM products").fetchall()
    }
    payload_prod_ids = {p["id"] for p in data.get("products", [])}

    # Delete current-partition products that the user removed
    if payload_prod_ids:
        ph = ",".join("?" * len(payload_prod_ids))
        conn.execute(
            f"DELETE FROM products WHERE partition=? AND id NOT IN ({ph})",
            [part] + list(payload_prod_ids)
        )
    else:
        conn.execute("DELETE FROM products WHERE partition=?", (part,))

    # Upsert products from payload (preserve 'both' partition for seed/migrated data)
    for p in data.get("products", []):
        pid = p["id"]
        existing_p = existing_prod_parts.get(pid)
        final_part = PARTITION_BOTH if existing_p == PARTITION_BOTH else part
        conn.execute("""
            INSERT OR REPLACE INTO products
              (id,name,category,unit,purchase,sale,gst,stock,min_stock,sku,expiry,brand,hsn,desc,partition)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            pid, p.get("name",""), p.get("category",""), p.get("unit","Tablet"),
            float(p.get("purchase",0)), float(p.get("sale",0)), float(p.get("gst",12)),
            int(p.get("stock",0)), int(p.get("minStock",10)),
            p.get("sku",""), p.get("expiry",""), p.get("brand",""),
            p.get("hsn",""), p.get("desc",""),
            final_part,
        ))

    # ── MEDICINE DB — Stock-ins (partition-aware upsert) ────────
    existing_si_parts = {
        r["id"]: r["partition"]
        for r in conn.execute("SELECT id, partition FROM stock_ins").fetchall()
    }
    payload_si_ids = {si["id"] for si in data.get("stockIns", [])}

    if payload_si_ids:
        ph = ",".join("?" * len(payload_si_ids))
        conn.execute(
            f"DELETE FROM stock_ins WHERE partition=? AND id NOT IN ({ph})",
            [part] + list(payload_si_ids)
        )
    else:
        conn.execute("DELETE FROM stock_ins WHERE partition=?", (part,))

    for si in data.get("stockIns", []):
        sid = si["id"]
        existing_sp = existing_si_parts.get(sid)
        final_part  = PARTITION_BOTH if existing_sp == PARTITION_BOTH else part
        conn.execute("""
            INSERT OR REPLACE INTO stock_ins
              (id,date,product_id,product_name,qty,price,batch,expiry,supplier,invoice_no,notes,partition)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            sid, si.get("date",""), si.get("productId",""), si.get("productName",""),
            int(si.get("qty",0)), float(si.get("price",0)),
            si.get("batch",""), si.get("expiry",""), si.get("supplier",""),
            si.get("invoiceNo",""), si.get("notes",""),
            final_part,
        ))

    # ── SALES HISTORY DB — Bills (replace current partition's type) ─
    payload_bill_ids = {b["id"] for b in data.get("bills", [])}

    if bill_type == PARTITION_WS:
        btype_filter = "bill_store_type = 'wholesale'"
    else:
        btype_filter = "bill_store_type != 'wholesale'"

    if payload_bill_ids:
        ph = ",".join("?" * len(payload_bill_ids))
        # Delete orphaned bill_items first
        conn.execute(f"""
            DELETE FROM bill_items WHERE bill_id IN (
                SELECT id FROM bills WHERE {btype_filter} AND id NOT IN ({ph})
            )""", list(payload_bill_ids))
        conn.execute(
            f"DELETE FROM bills WHERE {btype_filter} AND id NOT IN ({ph})",
            list(payload_bill_ids)
        )
    else:
        conn.execute(f"DELETE FROM bill_items WHERE bill_id IN (SELECT id FROM bills WHERE {btype_filter})")
        conn.execute(f"DELETE FROM bills WHERE {btype_filter}")

    for b in data.get("bills", []):
        # Only sync bills belonging to current partition
        b_type = b.get("billStoreType", "retail")
        if (part == PARTITION_WS and b_type != "wholesale") or \
           (part == PARTITION_RT and b_type == "wholesale"):
            continue

        conn.execute("""
            INSERT OR REPLACE INTO bills
              (id,bill_no,date,customer,phone,doctor,rx,payment_mode,notes,
               subtotal,total_discount,total_gst,round_off,grand_total,bill_store_type,
               ws_supplier,ws_owner,ws_gstin,shop_name,shopkeeper_gstin,
               rt_shop,rt_owner,rt_gstin,rt_license,rt_email,rt_phone)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            b["id"], b.get("billNo",""), b.get("date",""),
            b.get("customer",""), b.get("phone",""), b.get("doctor",""), b.get("rx",""),
            b.get("paymentMode","Cash"), b.get("notes",""),
            float(b.get("subtotal",0)), float(b.get("totalDiscount",0)),
            float(b.get("totalGst",0)), float(b.get("roundOff",0)), float(b.get("grandTotal",0)),
            b.get("billStoreType","retail"),
            b.get("wsSupplier",""), b.get("wsOwner",""), b.get("wsGstin",""),
            b.get("shopName",""), b.get("shopkeeperGstin",""),
            b.get("rtShop",""), b.get("rtOwner",""), b.get("rtGstin",""),
            b.get("rtLicense",""), b.get("rtEmail",""), b.get("rtPhone",""),
        ))
        for it in b.get("items", []):
            ga, lt = _calc_item(
                float(it.get("unitPrice",0)), float(it.get("qty",0)),
                float(it.get("discount",0)), float(it.get("gstRate",0)))
            conn.execute("""
                INSERT OR REPLACE INTO bill_items
                  (id,bill_id,product_id,name,category,unit,qty,unit_price,
                   discount,gst_rate,gst_amt,line_total)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                it.get("id") or uid(), b["id"],
                it.get("productId",""), it.get("name",""),
                it.get("category",""), it.get("unit",""),
                float(it.get("qty",0)), float(it.get("unitPrice",0)),
                float(it.get("discount",0)), float(it.get("gstRate",0)),
                float(it.get("gstAmt", ga)), float(it.get("lineTotal", lt)),
            ))

    # ── CREDIT DB — Wholesale credits (full replace) ───────────
    conn.execute("DELETE FROM credits")
    for c in data.get("credits", []):
        conn.execute("""
            INSERT INTO credits
              (id,date,shop_name,shopkeeper_name,phone,for_item,amount,method,status,partition)
            VALUES(?,?,?,?,?,?,?,?,?,?)
        """, (
            c["id"], c.get("date",""), c.get("shopName",""), c.get("shopkeeperName",""),
            c.get("phone",""), c.get("forItem",""),
            float(c.get("amount",0)),
            c.get("method","Cash"), c.get("status","Pending"),
            c.get("partition", PARTITION_WS),
        ))

    # ── CREDIT DB — Retail shop_credits (full replace) ─────────
    conn.execute("DELETE FROM shop_credits")
    for sc in data.get("shopCredits", []):
        pending = sc.get("pending", max(0, sc.get("totalPurchase",0) - sc.get("paid",0)))
        conn.execute("""
            INSERT INTO shop_credits
              (id,supplier_id,supplier_name,owner_name,total_purchase,paid,
               payment_mode,pending,last_purchase_date,bill_date,status,partition)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            sc["id"], sc.get("supplierId",""), sc.get("supplierName",""), sc.get("ownerName",""),
            float(sc.get("totalPurchase",0)), float(sc.get("paid",0)),
            sc.get("paymentMode","Cash"), float(pending),
            sc.get("lastPurchaseDate",""), sc.get("billDate",""),
            sc.get("status","Pending"),
            sc.get("partition", PARTITION_RT),
        ))

    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────
# IMPORT  (migrate from localStorage STATE JSON)
# All imported records are tagged partition='both' so they are
# visible regardless of which pharmacy mode is active.
# ─────────────────────────────────────────────────────────────
@app.route('/api/import', methods=['POST'])
@require_json
def import_data():
    """
    One-shot migration from the original localStorage STATE object.
    POST the JSON from localStorage key 'pharmacare_v2'.
    Safe to call multiple times — uses INSERT OR IGNORE.
    All imported records are tagged partition='both' (visible in all modes).
    """
    data = request.get_json()
    conn = get_db()

    # Settings
    s = data.get("settings", {})
    conn.execute("""
        UPDATE settings SET
          store_name=?, store_type=?, address=?, phone=?, email=?,
          license_no=?, gstin=?, default_gst=?, currency=?,
          low_stock_threshold=?, expiry_alert_days=?,
          supplier_name=?, wholesaler=?, owner_name=?, wholesaler_id=?,
          shop_name=?, retailer_owner=?,
          wholesale_upi_qr=?, retail_upi_qr=?, next_bill_no=?
        WHERE id=1
    """, (
        s.get("storeName",         "My Pharmacy"),
        s.get("storeType",         "Retail Pharmacy"),
        s.get("address",           ""),
        s.get("phone",             ""),
        s.get("email",             ""),
        s.get("license",           ""),
        s.get("gstin",             ""),
        float(s.get("defaultGst",  12)),
        s.get("currency",          "₹"),
        int(s.get("lowStockThreshold", 10)),
        int(s.get("expiryAlertDays",   90)),
        s.get("supplierName",      ""),
        s.get("wholesaler",        ""),
        s.get("ownerName",         ""),
        s.get("wholesalerId",      ""),
        s.get("shopName",          ""),
        s.get("retailerOwner",     ""),
        s.get("wholesaleUpiQr",    ""),
        s.get("retailUpiQr",       ""),
        int(data.get("nextBillNo", s.get("nextBillNo", 1))),
    ))

    for c in data.get("categories", []):
        conn.execute(
            "INSERT OR IGNORE INTO categories(id,name,desc) VALUES(?,?,?)",
            (c["id"], c["name"], c.get("desc", ""))
        )

    # All migrated products → partition='both' (visible in all modes)
    for p in data.get("products", []):
        conn.execute("""
            INSERT OR IGNORE INTO products
              (id,name,category,unit,purchase,sale,gst,stock,min_stock,sku,expiry,brand,hsn,desc,partition)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            p["id"], p["name"],
            p.get("category", ""),
            p.get("unit",     "Tablet"),
            float(p.get("purchase",  0)),
            float(p.get("sale",      0)),
            float(p.get("gst",      12)),
            int(p.get("stock",       0)),
            int(p.get("minStock",   10)),
            p.get("sku",    ""),
            p.get("expiry", ""),
            p.get("brand",  ""),
            p.get("hsn",    ""),
            p.get("desc",   ""),
            PARTITION_BOTH,     # ← visible to all modes
        ))

    # All migrated stock-ins → partition='both'
    for si in data.get("stockIns", []):
        conn.execute("""
            INSERT OR IGNORE INTO stock_ins
              (id,date,product_id,product_name,qty,price,batch,expiry,supplier,invoice_no,notes,partition)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            si["id"], si.get("date", ""),
            si.get("productId", ""), si.get("productName", ""),
            int(si.get("qty", 0)), float(si.get("price", 0)),
            si.get("batch", ""), si.get("expiry", ""),
            si.get("supplier", ""), si.get("invoiceNo", ""), si.get("notes", ""),
            PARTITION_BOTH,     # ← visible to all modes
        ))

    for b in data.get("bills", []):
        conn.execute("""
            INSERT OR IGNORE INTO bills
              (id,bill_no,date,customer,phone,doctor,rx,payment_mode,notes,
               subtotal,total_discount,total_gst,round_off,grand_total,bill_store_type,
               ws_supplier,ws_owner,ws_gstin,shop_name,shopkeeper_gstin,
               rt_shop,rt_owner,rt_gstin,rt_license,rt_email,rt_phone)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            b["id"], b.get("billNo", ""), b.get("date", ""),
            b.get("customer", ""), b.get("phone", ""),
            b.get("doctor", ""), b.get("rx", ""),
            b.get("paymentMode", "Cash"), b.get("notes", ""),
            float(b.get("subtotal",      0)),
            float(b.get("totalDiscount", 0)),
            float(b.get("totalGst",      0)),
            float(b.get("roundOff",      0)),
            float(b.get("grandTotal",    0)),
            b.get("billStoreType", "retail"),
            b.get("wsSupplier", ""), b.get("wsOwner", ""),
            b.get("wsGstin", ""),   b.get("shopName", ""),
            b.get("shopkeeperGstin", ""),
            b.get("rtShop", ""),    b.get("rtOwner", ""),
            b.get("rtGstin", ""),   b.get("rtLicense", ""),
            b.get("rtEmail", ""),   b.get("rtPhone", ""),
        ))
        for it in b.get("items", []):
            gst_amt, line_total = _calc_item(
                float(it.get("unitPrice", 0)), float(it.get("qty", 0)),
                float(it.get("discount",  0)), float(it.get("gstRate", 0)),
            )
            conn.execute("""
                INSERT OR IGNORE INTO bill_items
                  (id,bill_id,product_id,name,category,unit,qty,unit_price,
                   discount,gst_rate,gst_amt,line_total)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                it.get("id") or uid(), b["id"],
                it.get("productId", ""), it.get("name", ""),
                it.get("category",  ""), it.get("unit",  ""),
                float(it.get("qty",       0)),
                float(it.get("unitPrice", 0)),
                float(it.get("discount",  0)),
                float(it.get("gstRate",   0)),
                float(it.get("gstAmt",    gst_amt)),
                float(it.get("lineTotal", line_total)),
            ))

    # Migrated wholesale credits → partition='both'
    for c in data.get("credits", []):
        conn.execute("""
            INSERT OR IGNORE INTO credits
              (id,date,shop_name,shopkeeper_name,phone,for_item,amount,method,status,partition)
            VALUES(?,?,?,?,?,?,?,?,?,?)
        """, (
            c["id"], c.get("date", ""),
            c.get("shopName", ""), c.get("shopkeeperName", ""),
            c.get("phone", ""), c.get("forItem", ""),
            float(c.get("amount", 0)),
            c.get("method", "Cash"), c.get("status", "Pending"),
            PARTITION_BOTH,     # ← visible to all modes
        ))

    # Migrated retail shop_credits → partition='both'
    for sc in data.get("shopCredits", []):
        pending = sc.get("pending", max(0, sc.get("totalPurchase", 0) - sc.get("paid", 0)))
        conn.execute("""
            INSERT OR IGNORE INTO shop_credits
              (id,supplier_id,supplier_name,owner_name,total_purchase,paid,
               payment_mode,pending,last_purchase_date,bill_date,status,partition)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            sc["id"],
            sc.get("supplierId", ""), sc.get("supplierName", ""),
            sc.get("ownerName", ""),
            float(sc.get("totalPurchase", 0)),
            float(sc.get("paid",          0)),
            sc.get("paymentMode", "Cash"),
            float(pending),
            sc.get("lastPurchaseDate", ""),
            sc.get("billDate", ""),
            sc.get("status", "Pending"),
            PARTITION_BOTH,     # ← visible to all modes
        ))

    conn.commit()
    counts = {t: conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
              for t in ["categories","products","bills","stock_ins","credits","shop_credits"]}
    conn.close()
    return jsonify({"ok": True, "message": "Data imported successfully (partition=both)", "counts": counts})


# ─────────────────────────────────────────────────────────────
# FULL BACKUP EXPORT  (all partitions — for complete backup)
# ─────────────────────────────────────────────────────────────
@app.route('/api/export/backup', methods=['GET'])
def export_full_backup():
    """Return ALL data from every partition as a single JSON backup."""
    conn = get_db()
    s    = _get_settings(conn)

    categories = rows(conn.execute("SELECT * FROM categories ORDER BY name").fetchall())
    products   = [_product_out(r) for r in conn.execute("SELECT * FROM products ORDER BY name").fetchall()]

    # All bills (both partitions)
    all_bills = []
    for b in conn.execute("SELECT * FROM bills ORDER BY date DESC").fetchall():
        items = rows(conn.execute("SELECT * FROM bill_items WHERE bill_id=?", (b["id"],)).fetchall())
        bill_dict = dict(b)
        bill_dict["items"] = [_bill_item_out(i) for i in conn.execute(
            "SELECT * FROM bill_items WHERE bill_id=?", (b["id"],)).fetchall()]
        all_bills.append({
            "id": b["id"], "billNo": b["bill_no"] or "", "date": b["date"] or "",
            "customer": b["customer"] or "", "phone": b["phone"] or "",
            "doctor": b["doctor"] or "", "rx": b["rx"] or "",
            "paymentMode": b["payment_mode"] or "Cash", "notes": b["notes"] or "",
            "subtotal": b["subtotal"] or 0, "totalDiscount": b["total_discount"] or 0,
            "totalGst": b["total_gst"] or 0, "roundOff": b["round_off"] or 0,
            "grandTotal": b["grand_total"] or 0,
            "billStoreType": b["bill_store_type"] or "retail",
            "wsSupplier": b["ws_supplier"] or "", "wsOwner": b["ws_owner"] or "",
            "wsGstin": b["ws_gstin"] or "", "shopName": b["shop_name"] or "",
            "shopkeeperGstin": b["shopkeeper_gstin"] or "",
            "rtShop": b["rt_shop"] or "", "rtOwner": b["rt_owner"] or "",
            "rtGstin": b["rt_gstin"] or "", "rtLicense": b["rt_license"] or "",
            "rtEmail": b["rt_email"] or "", "rtPhone": b["rt_phone"] or "",
            "items": [_bill_item_out(i) for i in conn.execute(
                "SELECT * FROM bill_items WHERE bill_id=?", (b["id"],)).fetchall()],
        })

    stock_ins    = [{"id": r["id"], "date": r["date"] or "", "productId": r["product_id"] or "",
                     "productName": r["product_name"] or "", "qty": r["qty"] or 0,
                     "price": r["price"] or 0, "batch": r["batch"] or "",
                     "expiry": r["expiry"] or "", "supplier": r["supplier"] or "",
                     "invoiceNo": r["invoice_no"] or "", "notes": r["notes"] or "",
                     "partition": r["partition"] if "partition" in r.keys() else PARTITION_BOTH}
                    for r in conn.execute("SELECT * FROM stock_ins ORDER BY date DESC").fetchall()]
    credits      = [_credit_out(r) for r in conn.execute("SELECT * FROM credits ORDER BY date DESC").fetchall()]
    shop_credits = [_shop_credit_out(r) for r in conn.execute("SELECT * FROM shop_credits ORDER BY bill_date DESC").fetchall()]

    reset_rows = conn.execute("SELECT store_type_key, reset_date FROM dashboard_resets").fetchall()
    dashboard_resets = {r["store_type_key"]: r["reset_date"] for r in reset_rows}
    conn.close()

    return jsonify({
        "settings":        _settings_out(s),
        "categories":      categories,
        "products":        products,
        "bills":           all_bills,
        "stockIns":        stock_ins,
        "credits":         credits,
        "shopCredits":     shop_credits,
        "nextBillNo":      s.get("next_bill_no", 1),
        "dashboardResets": dashboard_resets,
        "_exportMeta": {
            "exportedAt": date.today().isoformat(),
            "source":     "PharmaCare Pro Backup",
            "version":    "2.0",
        }
    })


# ─────────────────────────────────────────────────────────────
# IMPORT — Medicines only  (CSV / XLSX pre-parsed to JSON)
# Partition assigned based on current pharmacy type in settings.
# ─────────────────────────────────────────────────────────────
@app.route('/api/import/medicines', methods=['POST'])
@require_json
def import_medicines():
    """
    Import medicine records from a JSON array.
    Frontend parses CSV/XLSX to JSON, sends here.
    Expected body: { "medicines": [ { name, category, unit, purchase, sale, gst,
                                       stock, minStock, sku, expiry, brand, hsn, desc }, ... ] }
    Partition is set from the current pharmacy type in settings.
    """
    data     = request.get_json()
    medicines = data.get("medicines", [])
    if not medicines:
        return jsonify({"error": "No medicine data found in the uploaded file"}), 400

    conn = get_db()
    part = _current_partition(conn)

    inserted = 0
    updated  = 0
    skipped  = 0
    errors   = []

    for m in medicines:
        name = (m.get("name") or "").strip()
        if not name:
            skipped += 1
            continue

        # Resolve category — accept name or id
        cat_raw = (m.get("category") or "").strip()
        cat_id  = None
        if cat_raw:
            # Try as ID first
            row_c = conn.execute("SELECT id FROM categories WHERE id=?", (cat_raw,)).fetchone()
            if row_c:
                cat_id = row_c["id"]
            else:
                # Try as name (case-insensitive)
                row_c = conn.execute(
                    "SELECT id FROM categories WHERE LOWER(name)=?", (cat_raw.lower(),)
                ).fetchone()
                if row_c:
                    cat_id = row_c["id"]
                else:
                    # Create the category
                    cat_id = uid()
                    conn.execute(
                        "INSERT OR IGNORE INTO categories(id,name,desc) VALUES(?,?,?)",
                        (cat_id, cat_raw, "")
                    )

        pid = m.get("id") or uid()

        # Check if already exists
        existing = conn.execute("SELECT id FROM products WHERE id=?", (pid,)).fetchone()
        if not existing:
            # Also check by name to avoid exact duplicates
            existing_by_name = conn.execute(
                "SELECT id FROM products WHERE LOWER(name)=? AND partition=?",
                (name.lower(), part)
            ).fetchone()
            if existing_by_name:
                pid = existing_by_name["id"]

        try:
            conn.execute("""
                INSERT OR REPLACE INTO products
                  (id,name,category,unit,purchase,sale,gst,stock,min_stock,
                   sku,expiry,brand,hsn,desc,partition)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                pid, name, cat_id,
                (m.get("unit") or m.get("form") or "Tablet"),
                float(m.get("purchase") or m.get("pur") or 0),
                float(m.get("sale") or m.get("mrp") or 0),
                float(m.get("gst") or 12),
                int(m.get("stock") or m.get("stk") or 0),
                int(m.get("minStock") or m.get("min_stock") or m.get("ms") or 10),
                (m.get("sku") or m.get("batch") or m.get("bat") or ""),
                (m.get("expiry") or m.get("exp") or ""),
                (m.get("brand") or m.get("manufacturer") or m.get("mfr") or ""),
                (m.get("hsn") or ""),
                (m.get("desc") or m.get("composition") or m.get("cmp") or ""),
                part,
            ))
            if existing:
                updated += 1
            else:
                inserted += 1
        except Exception as ex:
            errors.append({"name": name, "error": str(ex)})
            skipped += 1

    conn.commit()
    conn.close()
    return jsonify({
        "ok": True,
        "partition": part,
        "message": f"Medicines imported into '{part}' partition",
        "inserted": inserted, "updated": updated, "skipped": skipped,
        "errors":  errors[:10],  # return first 10 errors only
    })


# ─────────────────────────────────────────────────────────────
# IMPORT — Sales History  (CSV / XLSX pre-parsed to JSON)
# ─────────────────────────────────────────────────────────────
@app.route('/api/import/sales-history', methods=['POST'])
@require_json
def import_sales_history():
    """
    Import sales/bill records.
    Expected body: { "bills": [ { billNo, date, customer, phone, doctor,
                                   paymentMode, grandTotal, subtotal, ... }, ... ] }
    bill_store_type is set from current partition.
    """
    data  = request.get_json()
    bills = data.get("bills", [])
    if not bills:
        return jsonify({"error": "No sales data found in the uploaded file"}), 400

    conn      = get_db()
    part      = _current_partition(conn)
    bill_type = PARTITION_WS if part == PARTITION_WS else PARTITION_RT

    inserted = 0
    skipped  = 0
    errors   = []

    for b in bills:
        bill_date = (b.get("date") or b.get("Date") or today_str())
        customer  = (b.get("customer") or b.get("Customer") or b.get("customerName") or "").strip()
        grand     = float(b.get("grandTotal") or b.get("grand_total") or b.get("GrandTotal") or 0)
        subtotal  = float(b.get("subtotal") or b.get("Subtotal") or grand)
        bid       = b.get("id") or uid()

        existing = conn.execute("SELECT id FROM bills WHERE id=?", (bid,)).fetchone()
        if existing:
            skipped += 1
            continue

        # Auto-increment bill number
        s  = _get_settings(conn)
        n, bn = _next_bill_no(conn)
        conn.execute("UPDATE settings SET next_bill_no=? WHERE id=1", (n + 1,))

        try:
            conn.execute("""
                INSERT OR IGNORE INTO bills
                  (id,bill_no,date,customer,phone,doctor,payment_mode,
                   subtotal,total_discount,total_gst,round_off,grand_total,bill_store_type)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                bid,
                b.get("billNo") or b.get("bill_no") or bn,
                bill_date, customer,
                (b.get("phone") or ""),
                (b.get("doctor") or ""),
                (b.get("paymentMode") or b.get("payment_mode") or b.get("PaymentMode") or "Cash"),
                subtotal,
                float(b.get("totalDiscount") or b.get("total_discount") or 0),
                float(b.get("totalGst") or b.get("total_gst") or 0),
                float(b.get("roundOff") or b.get("round_off") or 0),
                grand,
                bill_type,
            ))
            inserted += 1
        except Exception as ex:
            errors.append({"customer": customer, "error": str(ex)})
            skipped += 1

    conn.commit()
    conn.close()
    return jsonify({
        "ok": True,
        "partition": part,
        "message": f"Sales history imported into '{bill_type}' store type",
        "inserted": inserted, "skipped": skipped,
        "errors": errors[:10],
    })


# ─────────────────────────────────────────────────────────────
# IMPORT — Credits  (CSV / XLSX pre-parsed to JSON)
# ─────────────────────────────────────────────────────────────
@app.route('/api/import/credits', methods=['POST'])
@require_json
def import_credits():
    """
    Import credit records.
    Wholesale: credits table. Retail/Hospital/Medical/Ayurvedic: shop_credits table.
    """
    data    = request.get_json()
    credits_data = data.get("credits", [])
    if not credits_data:
        return jsonify({"error": "No credit data found in the uploaded file"}), 400

    conn = get_db()
    part = _current_partition(conn)

    inserted = 0
    skipped  = 0
    errors   = []

    if part == PARTITION_WS:
        # Insert into wholesale credits table
        for c in credits_data:
            shop = (c.get("shopName") or c.get("shop_name") or c.get("ShopName") or "").strip()
            if not shop:
                skipped += 1
                continue
            cid = c.get("id") or uid()
            existing = conn.execute("SELECT id FROM credits WHERE id=?", (cid,)).fetchone()
            if existing:
                skipped += 1
                continue
            try:
                conn.execute("""
                    INSERT OR IGNORE INTO credits
                      (id,date,shop_name,shopkeeper_name,phone,for_item,amount,method,status,partition)
                    VALUES(?,?,?,?,?,?,?,?,?,?)
                """, (
                    cid,
                    (c.get("date") or c.get("Date") or today_str()),
                    shop,
                    (c.get("shopkeeperName") or c.get("shopkeeper_name") or c.get("OwnerName") or ""),
                    (c.get("phone") or ""),
                    (c.get("forItem") or c.get("for_item") or c.get("Item") or ""),
                    float(c.get("amount") or c.get("Amount") or 0),
                    (c.get("method") or c.get("Method") or "Cash"),
                    (c.get("status") or "Pending"),
                    PARTITION_WS,
                ))
                inserted += 1
            except Exception as ex:
                errors.append({"shop": shop, "error": str(ex)})
                skipped += 1
    else:
        # Insert into retail shop_credits table
        for sc in credits_data:
            supplier = (sc.get("supplierName") or sc.get("supplier_name") or sc.get("SupplierName") or "").strip()
            if not supplier:
                skipped += 1
                continue
            scid = sc.get("id") or uid()
            existing = conn.execute("SELECT id FROM shop_credits WHERE id=?", (scid,)).fetchone()
            if existing:
                skipped += 1
                continue
            try:
                total  = float(sc.get("totalPurchase") or sc.get("total_purchase") or 0)
                paid   = float(sc.get("paid") or 0)
                pending = float(sc.get("pending") or max(0, total - paid))
                conn.execute("""
                    INSERT OR IGNORE INTO shop_credits
                      (id,supplier_id,supplier_name,owner_name,total_purchase,paid,
                       payment_mode,pending,last_purchase_date,bill_date,status,partition)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    scid,
                    (sc.get("supplierId") or sc.get("supplier_id") or ""),
                    supplier,
                    (sc.get("ownerName") or sc.get("owner_name") or ""),
                    total, paid,
                    (sc.get("paymentMode") or sc.get("payment_mode") or "Cash"),
                    pending,
                    (sc.get("lastPurchaseDate") or sc.get("last_purchase_date") or today_str()),
                    (sc.get("billDate") or sc.get("bill_date") or today_str()),
                    (sc.get("status") or ("Cleared" if pending <= 0 else "Pending")),
                    PARTITION_RT,
                ))
                inserted += 1
            except Exception as ex:
                errors.append({"supplier": supplier, "error": str(ex)})
                skipped += 1

    conn.commit()
    conn.close()
    return jsonify({
        "ok": True,
        "partition": part,
        "message": f"Credits imported into '{part}' partition",
        "inserted": inserted, "skipped": skipped,
        "errors": errors[:10],
    })


# ─────────────────────────────────────────────────────────────
# SEED EXCEL MEDICINES  — loads med_query__1_.xlsx on first run
# All records tagged partition='both' (visible in ALL modes).
# Column mapping:
#   Column1.id   → id          Column1.name → name
#   Column1.cid  → category    Column1.form → unit
#   Column1.pur  → purchase    Column1.mrp  → sale
#   Column1.gst  → gst         Column1.stk  → stock
#   Column1.ms   → min_stock   Column1.bat  → sku (batch)
#   Column1.exp  → expiry      Column1.mfr  → brand
#   Column1.hsn  → hsn         Column1.cmp  → desc
# ─────────────────────────────────────────────────────────────
def seed_excel_medicines():
    """
    Load medicines from med_query__1_.xlsx into products table.
    Safe to call multiple times — uses INSERT OR IGNORE.
    All records tagged partition='both' → visible in ALL pharmacy modes.
    """
    xlsx_path = os.path.join(BASE_DIR, 'med_query__1_.xlsx')
    if not os.path.exists(xlsx_path):
        print("  ℹ  med_query__1_.xlsx not found — skipping Excel medicine seed")
        return

    try:
        import csv, io
        # Try pandas first, fall back to openpyxl
        try:
            import pandas as pd
            df = pd.read_excel(xlsx_path)
            records = df.to_dict('records')
        except ImportError:
            from openpyxl import load_workbook
            wb  = load_workbook(xlsx_path, read_only=True, data_only=True)
            ws  = wb.active
            hdr = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(max_row=1))]
            records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                records.append(dict(zip(hdr, row)))
    except Exception as e:
        print(f"  ✗ Could not read med_query__1_.xlsx: {e}")
        return

    if not records:
        print("  ℹ  med_query__1_.xlsx is empty — skipping")
        return

    conn = get_db()

    # ── Build a map of category_id → category_id (they're already IDs in the sheet)
    # We'll insert categories on-the-fly if they don't exist.
    # The sheet uses opaque hex IDs for categories (Column1.cid).
    # Map known cid values to human-readable names from the dataset.
    cid_name_map = {
        '5f122f32': 'Analgesics',
        '198ee715': 'Antibiotics',
        '70dfa1e7': 'Antacids & GI',
        '892b6604': 'Antihistamines',
        'e995823e': 'Vitamins & Supplements',
        'fa48a4f5': 'Antidiabetics',
        '15c1fd0a': 'Cardiovascular',
        'b4681e22': 'Syrups & Liquids',
        '9048aed2': 'Topical',
        '30473ab8': 'Respiratory',
        '6c8109f5': 'Neurological',
        '79c2f85f': 'Hormonal',
        'b299e67d': 'Eye Care',
        'd8c361a7': 'Dental & Oral',
        '41b2800f': 'Surgical & Consumables',
    }

    inserted = 0
    skipped  = 0

    for rec in records:
        # Column names as they appear in the xlsx (with 'Column1.' prefix)
        def g(key, alt=''):
            v = rec.get(f'Column1.{key}')
            if v is None:
                v = rec.get(key)
            return v if v is not None else alt

        pid  = str(g('id',   '')).strip()
        name = str(g('name', '')).strip()
        if not pid or not name or name == 'nan':
            skipped += 1
            continue

        cid  = str(g('cid',  '')).strip()
        form = str(g('form', 'Tablet')).strip() or 'Tablet'
        try:    pur = float(g('pur', 0))
        except: pur = 0.0
        try:    mrp = float(g('mrp', 0))
        except: mrp = 0.0
        try:    gst = float(g('gst', 12))
        except: gst = 12.0
        try:    stk = int(float(g('stk', 0)))
        except: stk = 0
        try:    ms  = int(float(g('ms', 10)))
        except: ms  = 10
        bat  = str(g('bat', '')).strip()
        if bat == 'nan': bat = ''
        exp  = str(g('exp', '')).strip()
        if exp == 'nan': exp = ''
        mfr  = str(g('mfr', '')).strip()
        if mfr == 'nan': mfr = ''
        hsn  = str(g('hsn', '')).strip()
        if hsn == 'nan': hsn = ''
        cmp  = str(g('cmp', '')).strip()
        if cmp == 'nan': cmp = ''

        # Ensure category exists
        cat_id = cid if cid else None
        if cat_id:
            exists_cat = conn.execute("SELECT id FROM categories WHERE id=?", (cat_id,)).fetchone()
            if not exists_cat:
                cat_name = cid_name_map.get(cat_id, f'Category {cat_id[:6]}')
                conn.execute(
                    "INSERT OR IGNORE INTO categories(id,name,desc) VALUES(?,?,?)",
                    (cat_id, cat_name, '')
                )

        try:
            conn.execute("""
                INSERT OR IGNORE INTO products
                  (id,name,category,unit,purchase,sale,gst,stock,min_stock,
                   sku,expiry,brand,hsn,desc,partition)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (pid, name, cat_id, form, pur, mrp, gst, stk, ms,
                  bat, exp, mfr, hsn, cmp, PARTITION_BOTH))
            inserted += 1
        except Exception as e:
            skipped += 1

    conn.commit()
    conn.close()
    print(f"  ✓ Excel medicines seeded: {inserted} inserted, {skipped} skipped")
    print(f"    (partition='both' — visible in Wholesale AND all Retail modes)")


# ─────────────────────────────────────────────────────────────
# SEED DATA  (auto-runs on first launch if DB is empty)
# All seed records get partition='both' — visible in ALL modes.
# ─────────────────────────────────────────────────────────────
def seed_demo_data():
    conn  = get_db()
    # If bills already exist, demo data was already seeded
    count = conn.execute("SELECT COUNT(*) FROM bills").fetchone()[0]
    if count > 0:
        conn.close()
        return  # already has demo billing data

    def mo(n):
        d = date.today()
        m = d.month + n
        y = d.year + (m - 1) // 12
        m = ((m - 1) % 12) + 1
        return f"{y}-{m:02d}"

    cats = [
        (uid(), 'Analgesics',             'Pain relievers'),
        (uid(), 'Antibiotics',            'Antibacterial medicines'),
        (uid(), 'Antacids',               'Stomach & digestion'),
        (uid(), 'Antihistamines',         'Allergy medicines'),
        (uid(), 'Vitamins & Supplements', 'Nutritional supplements'),
        (uid(), 'Antidiabetics',          'Diabetes medicines'),
        (uid(), 'Cardiovascular',         'Heart & BP medicines'),
        (uid(), 'Syrups & Liquids',       'Liquid medicines'),
        (uid(), 'Topical',                'Creams, gels & ointments'),
    ]
    for c in cats:
        conn.execute("INSERT OR IGNORE INTO categories(id,name,desc) VALUES(?,?,?)", c)

    # All seed products → partition='both'
    prods = [
        (uid(),'Paracetamol 500mg',  0,'Tablet', 12, 22, 5,200, 50,'B240101',mo(18),'Cipla',         'Paracetamol 500mg',           '30049099'),
        (uid(),'Amoxicillin 250mg',  1,'Capsule',55, 85,12,  8, 20,'SP2024A',mo(10),'Sun Pharma',     'Amoxicillin trihydrate 250mg','30041090'),
        (uid(),'Azithromycin 500mg', 1,'Tablet', 78,120,12, 45, 20,'LU4422', mo(2), 'Lupin',          'Azithromycin 500mg',          '30041090'),
        (uid(),'Cetirizine 10mg',    3,'Tablet', 18, 35, 5,150, 30,'MK0001', mo(24),'Mankind Pharma', 'Cetirizine HCl 10mg',         '30049099'),
        (uid(),'Omeprazole 20mg',    2,'Capsule',30, 55,12,  4, 25,'DR7788', mo(8), "Dr. Reddy's",    'Omeprazole 20mg',             '30049099'),
        (uid(),'Cough Syrup 100ml',  7,'Bottle', 40, 65,12,  5, 15,'PF990X', mo(6), 'Pfizer',         'Dextromethorphan+Guaifenesin','30049039'),
        (uid(),'Metformin 500mg',    5,'Tablet', 20, 42, 5,300,100,'USV0055',mo(16),'USV Ltd',         'Metformin HCl 500mg',         '30049099'),
        (uid(),'Amlodipine 5mg',     6,'Tablet', 22, 38, 5,180, 50,'CF2200', mo(20),'Cadila',          'Amlodipine Besylate 5mg',     '30049099'),
        (uid(),'Vitamin C 500mg',    4,'Tablet', 14, 28, 5,500, 50,'HM8800', mo(30),'Himalaya',        'Ascorbic Acid 500mg',         '29362700'),
        (uid(),'Betadine Cream 10g', 8,'Cream',  28, 48,12, 60, 10,'WM3300', mo(22),'Win Medicare',    'Povidone-Iodine 5%',          '30049039'),
        (uid(),'Eye Drops 5ml',      0,'Drops',  55, 90,12, 12, 10,'AL2023', mo(-2),'Alcon',           'Moxifloxacin 0.5%',           '30049039'),
        (uid(),'Pantoprazole 40mg',  2,'Tablet', 25, 45,12,  0, 30,'SR4411', mo(14),'Serum',           'Pantoprazole Sodium 40mg',    '30049099'),
    ]
    for p in prods:
        pid, name, ci, unit, pur, sal, gst, stk, mns, sku, exp, brand, desc, hsn = p
        cat_id = cats[ci][0]
        conn.execute("""
            INSERT OR IGNORE INTO products
              (id,name,category,unit,purchase,sale,gst,stock,min_stock,sku,expiry,brand,desc,hsn,partition)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (pid, name, cat_id, unit, pur, sal, gst, stk, mns, sku, exp, brand, desc, hsn,
              PARTITION_BOTH))

    # Demo bills — tagged retail, partition_both handled by bill_store_type
    patients = ['Ramesh Kumar','Priya Sharma','Anil Patel','Sunita Rao',
                'Vijay Singh','Meena Joshi','Deepak Nair']
    doctors  = ['Dr. Mehta','Dr. Singh','Dr. Verma','Dr. Pillai','Dr. Khan']
    modes    = ['Cash','UPI','Cash','Card','UPI','Cash','Insurance','Cash']

    for i in range(14):
        bill_date = (date.today() - timedelta(days=i % 9)).isoformat()
        p1 = prods[i % len(prods)]
        p2 = prods[(i + 3) % len(prods)]

        def make_it(p, qty, disc):
            pur, sal, gst = p[4], p[5], p[6]
            lt  = qty * sal
            da  = lt * disc / 100
            tax = lt - da
            ga  = round(tax * gst / 100, 2)
            ltt = round(tax + ga, 2)
            return (uid(), p[0], p[1], cats[p[2]][0], p[3],
                    qty, sal, disc, gst, ga, ltt)

        items = [make_it(p1, (i%5)+2, 0), make_it(p2, (i%3)+1, 5 if i%3==0 else 0)]
        sub   = sum(it[5]*it[6] for it in items)
        td    = sum(it[5]*it[6]*it[7]/100 for it in items)
        tg    = sum(it[9] for it in items)
        raw   = sub - td + tg
        grand = round(raw)
        bid   = uid()
        bn    = str(i + 1).zfill(4)
        # Seed 12 retail + 2 wholesale demo bills
        b_type = 'wholesale' if i >= 12 else 'retail'
        conn.execute("""
            INSERT OR IGNORE INTO bills
              (id,bill_no,date,customer,phone,doctor,payment_mode,
               subtotal,total_discount,total_gst,round_off,grand_total,bill_store_type)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (bid, bn, bill_date,
              patients[i % len(patients)], '', doctors[i % len(doctors)],
              modes[i % len(modes)],
              round(sub,2), round(td,2), round(tg,2), round(grand-raw,2), grand, b_type))
        for it in items:
            conn.execute("""
                INSERT OR IGNORE INTO bill_items
                  (id,bill_id,product_id,name,category,unit,qty,unit_price,
                   discount,gst_rate,gst_amt,line_total)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
            """, (it[0], bid) + it[1:])

    conn.execute("UPDATE settings SET next_bill_no=15 WHERE id=1")

    # Demo stock-ins → partition='both'
    for i in range(6):
        p = prods[i * 2 % len(prods)]
        conn.execute("""
            INSERT OR IGNORE INTO stock_ins
              (id,date,product_id,product_name,qty,price,batch,expiry,supplier,invoice_no,partition)
            VALUES(?,?,?,?,?,?,?,?,?,?,?)
        """, (uid(),
              (date.today()-timedelta(days=i+1)).isoformat(),
              p[0], p[1], 50+i*10, p[4], f"B{1000+i}", mo(12+i),
              'Main Distributor', f"INV-{100+i}",
              PARTITION_BOTH))

    # Demo wholesale credits → partition='both'
    c_shops = [
        ('Ramesh Medical Store',  'Ramesh Kumar',    '9876543210'),
        ('Priya Pharma Traders',  'Priya Sharma',    '9823456780'),
        ('Anil Drug House',       'Anil Patel',      '9912345678'),
        ('Sunita Medicals',       'Sunita Rao',      '9988776655'),
        ('Vijay Health Store',    'Vijay Singh',     '9765432109'),
        ('Meena Pharmaceuticals', 'Meena Joshi',     '9654321098'),
        ('Deepak Medical Agency', 'Deepak Nair',     '9543210987'),
        ('Kumar Drug Centre',     'Suresh Kumar',    '9432109876'),
        ('Patel Pharma Dist.',    'Rakesh Patel',    '9321098765'),
        ('Singh Medicals',        'Harpreet Singh',  '9210987654'),
        ('Jain Medical Traders',  'Abhay Jain',      '9109876543'),
        ('Sharma Drug House',     'Mohan Sharma',    '9098765432'),
    ]
    c_items   = ['Paracetamol 500mg x100','Amoxicillin 250mg x50','Azithromycin 500mg x30',
                 'Cetirizine 10mg x200','Omeprazole 20mg x80','Metformin 500mg x150',
                 'Vitamin C 500mg x100','Cough Syrup 100ml x20','Amlodipine 5mg x120',
                 'Betadine Cream 10g x40','Eye Drops 5ml x60','Pantoprazole 40mg x90']
    c_methods = ['UPI','NEFT','Cash','Credit/Debit Card','UPI','NEFT',
                 'Cash','UPI','NEFT','Cash','Credit/Debit Card','UPI']
    c_status  = ['Pending','Cleared','Pending','Pending','Cleared','Pending',
                 'Cleared','Pending','Pending','Cleared','Pending','Cleared']
    c_amounts = [1850,3200,2750,4100,1500,5600,2300,3800,2100,4900,1700,3500]
    c_days    = [2,5,8,12,18,22,28,35,45,60,75,85]
    for i, cs in enumerate(c_shops):
        conn.execute("""
            INSERT OR IGNORE INTO credits
              (id,date,shop_name,shopkeeper_name,phone,for_item,amount,method,status,partition)
            VALUES(?,?,?,?,?,?,?,?,?,?)
        """, (uid(), (date.today()-timedelta(days=c_days[i])).isoformat(),
              cs[0], cs[1], cs[2], c_items[i], c_amounts[i], c_methods[i], c_status[i],
              PARTITION_BOTH))

    # Demo retail shop-credits → partition='both'
    suppliers = [
        ('Apex Pharma Dist.',   'WHL-001','Rajesh Gupta',  12500,10000),
        ('MedLine Wholesale',   'WHL-002','Sanjay Mehta',   8200, 8200),
        ('BharatMed Traders',   'WHL-003','Vikram Shah',   15000, 8000),
        ('Sunrise Drug House',  'WHL-004','Pooja Reddy',    6800, 6800),
        ('National Pharma Co.', 'WHL-005','Arvind Kumar',  22000,15000),
        ('HealthFirst Dist.',   'WHL-006','Suresh Nair',    9400, 5000),
        ('Prime Med Supply',    'WHL-007','Deepa Iyer',    11200, 7000),
        ('City Drug Traders',   'WHL-008','Ravi Sharma',    7600, 5000),
        ('Lifeline Wholesale',  'WHL-009','Anita Patel',   18500,12000),
        ('GreenMed Dist.',      'WHL-010','Kartik Joshi',   5300, 3000),
    ]
    s_methods = ['UPI','NEFT','Cash','UPI','NEFT','Cash','Credit/Debit Card','UPI','NEFT','Cash']
    s_days    = [3, 8,12,20,25,32,40,50,62,75]
    for i, sup in enumerate(suppliers):
        pending = round(sup[3] - sup[4], 2)
        d_back  = (date.today()-timedelta(days=s_days[i])).isoformat()
        conn.execute("""
            INSERT OR IGNORE INTO shop_credits
              (id,supplier_id,supplier_name,owner_name,total_purchase,paid,
               payment_mode,pending,last_purchase_date,bill_date,status,partition)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
        """, (uid(), sup[1], sup[0], sup[2], sup[3], sup[4],
              s_methods[i], pending, d_back, d_back,
              'Cleared' if pending <= 0 else 'Pending',
              PARTITION_BOTH))

    conn.commit()
    conn.close()
    print("  ✓ Demo data seeded (12 medicines, 14 bills, credits, shop-credits)")
    print("  ✓ All seed records tagged partition='both' (visible in all pharmacy modes)")


# ─────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────
if __name__ == '__main__':
    init_db()
    seed_excel_medicines()   # loads med_query__1_.xlsx medicines into DB (partition='both')
    seed_demo_data()         # seeds demo bills/credits if DB is empty
    print("=" * 62)
    print("  PharmaCare Pro — Flask + SQLite Backend (Partition Edition)")
    print("  Open in browser :  http://localhost:5000")
    print("  Database file   :  pharmacare.db")
    print()
    print("  ── Partition Architecture ─────────────────────────────")
    print("  Settings → Pharmacy Type → determines active partition")
    print()
    print("  Wholesale Pharma   → 'wholesale' partition")
    print("  Retail Pharmacy  ┐")
    print("  Hospital Pharmacy├→ 'retail'    partition")
    print("  Medical Store    │")
    print("  Ayurvedic Store  ┘")
    print()
    print("  partition='both'   → seed/migrated data (all modes)")
    print()
    print("  ── Sub-databases inside pharmacare.db ─────────────────")
    print("  Medicine DB      : products, stock_ins   (partitioned)")
    print("  Sales History DB : bills, bill_items     (partitioned)")
    print("  Credit DB WS     : credits               (wholesale)")
    print("  Credit DB RT     : shop_credits          (retail)")
    print("  ──────────────────────────────────────────────────────")
    print()
    print("  ── Migrate from old localStorage app ─────────────────")
    print("  1. Open OLD app in browser (file:// version)")
    print("  2. Open DevTools Console  (F12 → Console tab)")
    print("  3. Run:  copy(localStorage.getItem('pharmacare_v2'))")
    print("  4. Paste into a file called: state.json")
    print("  5. Run: python migrate.py")
    print("  (Migrated data will be tagged partition='both')")
    print("  ─────────────────────────────────────────────────────")
    print()
    app.run(debug=True, port=5000)
